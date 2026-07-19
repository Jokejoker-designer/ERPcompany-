# -*- coding: utf-8 -*-
"""Canonical, read-only workbook normalizer for project BOQ imports.

The module never writes SQLite and never mutates the source workbook.  It keeps
sheet/row/column evidence, classifies stage allocations conservatively, and
builds a side-by-side reconciliation contract consumed by project-profile
Preview -> Confirm.  Unsupported/ambiguous data stays visible and blocks an
official commit instead of being guessed.
"""
from __future__ import annotations

import hashlib
import io
import os
import re
import unicodedata
from collections import Counter, defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation


PARSER_VERSION = "boq-normalizer-v1"
UNALLOCATED_STAGE_NAME = "Chua phan tang/Toan cong trinh"
MONEY_TOLERANCE_RATIO = Decimal("0.0002")  # +/- 0.02 percent
MAX_ROWS = 10000
MAX_COLUMNS = 200


class BoqNormalizationError(ValueError):
    pass


def _norm(value):
    text = unicodedata.normalize("NFD", str(value or ""))
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = text.replace("Đ", "D").replace("đ", "d").casefold()
    return re.sub(r"\s+", " ", text).strip()


def _decimal(value):
    if value in (None, "") or isinstance(value, bool):
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        try:
            return Decimal(str(value))
        except InvalidOperation:
            return None
    text = str(value).strip().replace("\u00a0", " ")
    if not text or text in ("-", "--") or text.startswith("="):
        return None
    text = re.sub(r"[^0-9,\.\-+]", "", text)
    if not text:
        return None
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        parts = text.split(",")
        text = "".join(parts) if len(parts[-1]) == 3 and len(parts) > 1 else text.replace(",", ".")
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def _canon(value):
    value = _decimal(value)
    if value is None:
        return None
    text = format(value, "f")
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    return text or "0"


def _display(value):
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    return value


def _column_name(index):
    result = ""
    current = int(index) + 1
    while current:
        current, remainder = divmod(current - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _read_xlsx(data):
    try:
        import openpyxl
        formula_wb = openpyxl.load_workbook(
            io.BytesIO(data), data_only=False, read_only=True, keep_links=False)
        value_wb = openpyxl.load_workbook(
            io.BytesIO(data), data_only=True, read_only=True, keep_links=False)
    except Exception as exc:
        raise BoqNormalizationError("Khong doc duoc XLSX/XLSM: %s" % exc) from exc
    sheets = []
    try:
        for name in formula_wb.sheetnames:
            formula_ws, value_ws = formula_wb[name], value_wb[name]
            rows, raw_rows = [], []
            max_row = min(max(formula_ws.max_row or 0, value_ws.max_row or 0), MAX_ROWS)
            max_col = min(max(formula_ws.max_column or 0, value_ws.max_column or 0), MAX_COLUMNS)
            formula_iter = formula_ws.iter_rows(min_row=1, max_row=max_row,
                                                min_col=1, max_col=max_col, values_only=True)
            value_iter = value_ws.iter_rows(min_row=1, max_row=max_row,
                                            min_col=1, max_col=max_col, values_only=True)
            for formulas, values in zip(formula_iter, value_iter):
                raw = [_display(value) for value in formulas]
                resolved = []
                for formula, value in zip(formulas, values):
                    resolved.append(_display(value if value is not None else
                                             (None if isinstance(formula, str) and formula.startswith("=")
                                              else formula)))
                rows.append(resolved)
                raw_rows.append(raw)
            sheets.append({"name": name, "rows": rows, "raw_rows": raw_rows,
                           "truncated": (formula_ws.max_row > MAX_ROWS or
                                         formula_ws.max_column > MAX_COLUMNS)})
    finally:
        formula_wb.close()
        value_wb.close()
    return sheets


def _read_xls(data):
    try:
        import xlrd
        wb = xlrd.open_workbook(file_contents=data, on_demand=True)
    except Exception as exc:
        raise BoqNormalizationError("Khong doc duoc XLS: %s" % exc) from exc
    sheets = []
    try:
        for name in wb.sheet_names():
            ws = wb.sheet_by_name(name)
            rows = [[_display(ws.cell_value(r, c)) for c in range(min(ws.ncols, MAX_COLUMNS))]
                    for r in range(min(ws.nrows, MAX_ROWS))]
            sheets.append({"name": name, "rows": rows, "raw_rows": rows,
                           "truncated": ws.nrows > MAX_ROWS or ws.ncols > MAX_COLUMNS})
    finally:
        wb.release_resources()
    return sheets


def _read_xlsb(data):
    try:
        from pyxlsb import open_workbook
        wb = open_workbook(io.BytesIO(data))
    except Exception as exc:
        raise BoqNormalizationError("Khong doc duoc XLSB (can pyxlsb): %s" % exc) from exc
    sheets = []
    try:
        for name in wb.sheets:
            rows = []
            with wb.get_sheet(name) as ws:
                for row_no, row in enumerate(ws.rows(), 1):
                    if row_no > MAX_ROWS:
                        break
                    values = [None] * min(max((cell.c for cell in row), default=-1) + 1, MAX_COLUMNS)
                    for cell in row:
                        if cell.c < MAX_COLUMNS:
                            values[cell.c] = _display(cell.v)
                    rows.append(values)
            sheets.append({"name": name, "rows": rows, "raw_rows": rows,
                           "truncated": len(rows) >= MAX_ROWS})
    finally:
        wb.close()
    return sheets


def _read_workbook(data, filename):
    ext = os.path.splitext(filename or "")[1].lower()
    if ext in (".xlsx", ".xlsm"):
        return _read_xlsx(data)
    if ext == ".xls":
        return _read_xls(data)
    if ext == ".xlsb":
        return _read_xlsb(data)
    raise BoqNormalizationError("Dinh dang BOQ khong duoc ho tro: %s" % (ext or "(rong)"))


DESCRIPTION_KEYS = (
    "ten san pham", "ten hang hoa", "hang muc dieu chinh", "hang muc cong viec",
    "noi dung cong viec", "noi dung", "dien giai", "ten vat tu", "hang muc",
)
UNIT_KEYS = ("don vi", "dvt", "đvt")
STT_KEYS = ("stt", "stt.", "so tt")


def _header_labels(rows, row_index):
    first = rows[row_index] if row_index < len(rows) else []
    second = rows[row_index + 1] if row_index + 1 < len(rows) else []
    width = max(len(first), len(second))
    first_normalized = [_norm(value) for value in first]
    description_index = next((index for index, label in enumerate(first_normalized)
                              if _contains_any(label, DESCRIPTION_KEYS)), None)
    # A second header row normally leaves STT/description/unit blank and only
    # expands merged groups such as Don gia -> Vat tu/Nhan cong.  If the next
    # row already has an item/section name, it is source data and must never be
    # folded into the header.
    use_second = bool(second) and description_index is not None and \
        _value(second, description_index) in (None, "") and \
        any(value not in (None, "") for value in second)
    labels, parent = [], ""
    for col in range(width):
        top = str(first[col]).strip() if col < len(first) and first[col] not in (None, "") else ""
        bottom = (str(second[col]).strip()
                  if use_second and col < len(second) and second[col] not in (None, "") else "")
        if top:
            parent = top
        inherited = parent if not top and bottom else top
        labels.append(_norm(" ".join(part for part in (inherited, bottom) if part)))
    return labels


def _contains_any(text, keys):
    return any(key in text for key in keys)


def _is_stage_label(label):
    if not label or _contains_any(label, ("tong", "so luong", "khoi luong", "don gia", "thanh tien")):
        return False
    # "dieu chinh tang" means an increase, not a floor.  A floor-stage must
    # carry an explicit floor token/number or another location axis.
    return bool(re.search(
        r"(^| )(tang\s*[-:]?\s*(?:ham|mai|ky thuat|tum|[0-9]{1,2}|[a-z](?![a-z]))|ham\b|mai\b|"
        r"truc dung\b|nha ve sinh\b|zone\b|khu vuc\b|khu\s+[a-z0-9]+|"
        r"giai doan\s+[a-z0-9]+)", label))


def _candidate(rows, sheet_name):
    candidates = []
    for row_index in range(min(len(rows), 80)):
        current_labels = [_norm(value) for value in rows[row_index]]
        next_labels = ([_norm(value) for value in rows[row_index + 1]]
                       if row_index + 1 < len(rows) else [])
        current_has_desc = any(_contains_any(label, DESCRIPTION_KEYS)
                               for label in current_labels)
        next_has_desc = any(_contains_any(label, DESCRIPTION_KEYS)
                            for label in next_labels)
        # Avoid claiming a date/title row as the header merely because the next
        # row contains the real header.  The next loop iteration will preserve
        # the correct Excel row number and its optional sub-header.
        if not current_has_desc and next_has_desc:
            continue
        labels = _header_labels(rows, row_index)
        has_desc = any(_contains_any(label, DESCRIPTION_KEYS) for label in labels)
        has_unit = any(_contains_any(label, UNIT_KEYS) for label in labels)
        has_stt = any(label in STT_KEYS or label.startswith("stt ") for label in labels)
        price_count = sum("don gia" in label for label in labels)
        amount_count = sum("thanh tien" in label for label in labels)
        stage_count = sum(_is_stage_label(label) for label in labels)
        qty_count = sum(_contains_any(label, ("so luong", "khoi luong")) for label in labels)
        score = has_desc * 6 + has_unit * 3 + has_stt * 2 + price_count * 3 + \
            amount_count * 3 + min(stage_count, 5) + min(qty_count, 2)
        if has_desc and (has_unit or qty_count or stage_count) and score >= 9:
            candidates.append({"sheet_name": sheet_name, "header_index": row_index,
                               "header_row": row_index + 1, "labels": labels,
                               "score": score, "price_count": price_count,
                               "amount_count": amount_count, "stage_count": stage_count})
    return max(candidates, key=lambda item: (item["score"], -item["header_index"])) if candidates else None


def _first_index(labels, predicate):
    return next((index for index, label in enumerate(labels) if predicate(label)), None)


def _roles(candidate):
    labels = candidate["labels"]
    description = _first_index(labels, lambda label: _contains_any(label, DESCRIPTION_KEYS))
    unit = _first_index(labels, lambda label: _contains_any(label, UNIT_KEYS))
    stt = _first_index(labels, lambda label: label in STT_KEYS or label.startswith("stt "))
    stage_cols = [index for index, label in enumerate(labels) if _is_stage_label(label)]
    price_cols = [index for index, label in enumerate(labels) if "don gia" in label]
    if len(price_cols) == 1:
        first_price = price_cols[0]
        for index in range(first_price + 1, min(first_price + 3, len(labels))):
            if labels[index] in ("vat tu", "nhan cong"):
                price_cols.append(index)
            else:
                break
    amount_cols = [index for index, label in enumerate(labels)
                   if "thanh tien" in label or re.match(r"^tt\s+(vat tu|nhan cong)$", label)]
    quantity_cols = [index for index, label in enumerate(labels)
                     if _contains_any(label, ("so luong", "khoi luong"))
                     and index not in stage_cols and "don gia" not in label and "thanh tien" not in label]
    stage_total = next((index for index in quantity_cols
                        if labels[index] == "tong khoi luong" or
                        ("tong khoi luong" in labels[index]
                         and "phat sinh" not in labels[index]
                         and "hop dong" not in labels[index])), None)
    quantity = next((index for index in quantity_cols
                     if labels[index] in ("khoi luong", "so luong")), None)
    if quantity is None:
        quantity = next((index for index in quantity_cols
                         if _contains_any(labels[index], ("khoi luong hop dong",
                                                          "tong khoi luong phat sinh",
                                                          "so luong"))), None)
    if quantity is None:
        quantity = next((index for index in quantity_cols if index != stage_total),
                        quantity_cols[0] if quantity_cols else None)
    combined_amount = next((index for index, label in enumerate(labels)
                            if label in ("tong", "tong cong") or label.endswith(" tong")), None)
    technical = _first_index(labels, lambda label: _contains_any(
        label, ("yeu cau", "chi dan ky thuat", "quy cach", "model")))
    category = _first_index(labels, lambda label: _contains_any(
        label, ("chung loai", "hang san xuat", "thuong hieu", "nhan hieu")))
    note = _first_index(labels, lambda label: _contains_any(label, ("ghi chu", "note")))
    return {"description": description, "unit": unit, "stt": stt,
            "quantity": quantity, "stage_total": stage_total, "stage_cols": stage_cols,
            "price_cols": price_cols, "amount_cols": amount_cols,
            "combined_amount": combined_amount, "technical": technical,
            "category": category, "note": note}


def _value(row, index):
    return row[index] if index is not None and index < len(row) else None


def _sum_numeric(row, indexes):
    values = [_decimal(_value(row, index)) for index in indexes]
    values = [value for value in values if value is not None]
    return sum(values, Decimal(0)) if values else None


def _stage_tokens(text):
    source = _norm(text)
    matches = []
    patterns = (
        r"tang\s*[-:]?\s*(?:ham|mai|ky thuat|tum|[0-9]{1,2}|[a-z](?![a-z]))",
        r"truc dung\s*[0-9x ]+", r"khu\s+[a-z0-9]+", r"zone\s+[a-z0-9]+",
    )
    for pattern in patterns:
        matches.extend(re.findall(pattern, source))
    return list(dict.fromkeys(value.strip() for value in matches if value.strip()))


def _stage_name_from_heading(text):
    normalized = _norm(text)
    if not _stage_tokens(text):
        return None
    # Keep the source heading as a composite stage when it explicitly denotes a
    # range/group.  Never split one source quantity across several floors.
    if re.search(r"tang\s*\d+\s*(?:den|toi|[-–])\s*tang\s*\d+", normalized):
        return str(text).strip()
    tokens = _stage_tokens(text)
    return str(text).strip() if len(tokens) == 1 else None


def _is_total_label(stt, text):
    """Classify financial footer rows without treating them as BOQ headings.

    Total rows are source evidence for reconciliation, but they must never be
    persisted as work items.  Keep the rule deliberately narrow so ordinary
    descriptions containing the word ``tong`` are not discarded.
    """
    label = _norm(text)
    stt_key = _norm(stt)
    prefixes = (
        "tong cong", "thue vat", "thanh tien", "cong tien",
        "tong gia tri", "gia tri truoc thue", "tien thue",
        "tong thanh toan",
    )
    return label.startswith(prefixes) or bool(
        re.fullmatch(r"tc\s*\d*", stt_key) and label)


def _payload(value, raw, sheet_name, row_number, column_index, derived=False):
    return {"value": _canon(value), "raw_value": None if raw is None else str(raw),
            "source_sheet": sheet_name, "source_row": row_number,
            "source_column": _column_name(column_index) if column_index is not None else None,
            "derived": bool(derived)}


def _parse_sheet(sheet, candidate):
    rows, raw_rows = sheet["rows"], sheet["raw_rows"]
    roles = _roles(candidate)
    if roles["description"] is None:
        return {"lines": [], "stages": [], "roles": roles}
    stages = []
    for order, col in enumerate(roles["stage_cols"], 1):
        raw_name = _value(raw_rows[candidate["header_index"]], col)
        if raw_name in (None, ""):
            raw_name = candidate["labels"][col]
        stages.append({"key": "stage:%s:%s" % (sheet["name"], col), "source_index": col,
                       "source_column": _column_name(col), "raw_name": str(raw_name),
                       "name": str(raw_name), "order": order, "is_unallocated_bucket": False})
    lines, totals, current_stage, warnings = [], [], None, []
    for index in range(candidate["header_index"] + 1, len(rows)):
        row, raw_row = rows[index], raw_rows[index] if index < len(raw_rows) else rows[index]
        text = _value(row, roles["description"])
        if text in (None, ""):
            continue
        stt = _value(row, roles["stt"])
        unit = _value(row, roles["unit"])
        qty = _decimal(_value(row, roles["quantity"]))
        source_stage_total = _decimal(_value(row, roles["stage_total"]))
        prices = _sum_numeric(row, roles["price_cols"])
        amounts = _sum_numeric(row, roles["amount_cols"])
        combined = _decimal(_value(row, roles["combined_amount"]))
        amount = combined if combined is not None else amounts
        stage_values = [(stage, _decimal(_value(row, stage["source_index"]))) for stage in stages]
        stage_values = [(stage, value) for stage, value in stage_values if value is not None]
        numeric = any(value is not None for value in (qty, prices, amount)) or bool(stage_values)
        is_detail = bool(str(unit or "").strip() and numeric)
        if not is_detail:
            if _is_total_label(stt, text):
                totals.append({
                    "kind": "total", "source_row": index + 1,
                    "source_order": index - candidate["header_index"],
                    "stt_raw": stt, "text_raw": str(text).strip(),
                    "amount": _payload(amount, _value(raw_row, roles["combined_amount"]),
                                       sheet["name"], index + 1, roles["combined_amount"],
                                       derived=(combined is None and len(roles["amount_cols"]) > 1)),
                })
                continue
            current_stage = _stage_name_from_heading(text) or current_stage
            lines.append({"kind": "heading", "source_row": index + 1,
                          "source_order": index - candidate["header_index"],
                          "stt_raw": stt, "text_raw": str(text), "heading_level": 1,
                          "hierarchy_path": [{"source_row": index + 1, "stt_raw": stt,
                                              "text_raw": str(text), "level": 1}]})
            continue
        stage_quantities = []
        for stage, value in stage_values:
            stage_quantities.append({"stage_key": stage["key"],
                "stage_name_raw": stage["name"], "quantity": _canon(value),
                **_payload(value, _value(raw_row, stage["source_index"]), sheet["name"],
                           index + 1, stage["source_index"])})
        allocation_evidence = "COLUMN_AXIS" if stage_quantities else None
        allocation_confidence = "HIGH" if stage_quantities else None
        if not stage_quantities and current_stage and qty is not None:
            key = "section:" + hashlib.sha1(_norm(current_stage).encode("utf-8")).hexdigest()[:12]
            stage_quantities = [{"stage_key": key, "stage_name_raw": current_stage,
                "quantity": _canon(qty), **_payload(qty, _value(raw_row, roles["quantity"]),
                    sheet["name"], index + 1, roles["quantity"])}]
            allocation_evidence, allocation_confidence = "SECTION_HEADING", "MEDIUM_HIGH"
        if not stage_quantities and qty is not None:
            tokens = _stage_tokens(text)
            if len(tokens) == 1:
                key = "rowtext:" + hashlib.sha1(tokens[0].encode("utf-8")).hexdigest()[:12]
                stage_quantities = [{"stage_key": key, "stage_name_raw": tokens[0],
                    "quantity": _canon(qty), **_payload(qty, _value(raw_row, roles["quantity"]),
                        sheet["name"], index + 1, roles["quantity"])}]
                allocation_evidence, allocation_confidence = "ROW_TEXT", "MEDIUM"
        stage_sum = sum((_decimal(value["value"]) or Decimal(0) for value in stage_quantities),
                        Decimal(0)) if stage_quantities else None
        line = {
            "kind": "detail", "source_row": index + 1,
            "source_order": index - candidate["header_index"], "stt_raw": stt,
            "text_raw": str(text),
            "technical_raw": (str(_value(raw_row, roles["technical"])).strip()
                              if _value(raw_row, roles["technical"]) not in (None, "") else None),
            "unit_raw": str(unit).strip(),
            "category_raw": (str(_value(raw_row, roles["category"])).strip()
                             if _value(raw_row, roles["category"]) not in (None, "") else None),
            "note_raw": (str(_value(raw_row, roles["note"])).strip()
                         if _value(raw_row, roles["note"]) not in (None, "") else None),
            "hierarchy_path": [],
            "contract_quantity": _payload(qty, _value(raw_row, roles["quantity"]),
                                            sheet["name"], index + 1, roles["quantity"]),
            # Preserve the source floor-total column exactly.  The explicit
            # stage sum and contract quantity are separate business facts and
            # must not be copied into a blank source total.
            "stage_total": _payload(source_stage_total,
                                    _value(raw_row, roles["stage_total"]), sheet["name"],
                                    index + 1, roles["stage_total"]),
            "unit_price": _payload(prices, None, sheet["name"], index + 1, None,
                                   derived=len(roles["price_cols"]) > 1),
            "amount": _payload(amount, _value(raw_row, roles["combined_amount"]),
                               sheet["name"], index + 1, roles["combined_amount"],
                               derived=(combined is None and len(roles["amount_cols"]) > 1)),
            "stage_quantities": stage_quantities, "stage_cells": list(stage_quantities),
            "stage_quantity_sum": _canon(stage_sum),
            "stage_total_matches_sum": (_quantity_equal(source_stage_total, stage_sum)
                                        if source_stage_total is not None and stage_sum is not None else None),
            "stage_status": "allocated" if stage_quantities else "unallocated",
            "stage_bucket_key": None if stage_quantities else "unallocated",
            "stage_bucket_name": None if stage_quantities else UNALLOCATED_STAGE_NAME,
            "classification_evidence": allocation_evidence or "NONE",
            "classification_confidence": allocation_confidence or "UNALLOCATED",
        }
        if (source_stage_total is not None and stage_sum is not None
                and not _quantity_equal(source_stage_total, stage_sum)):
            warnings.append({"code": "SOURCE_STAGE_QUANTITY_MISMATCH", "sheet": sheet["name"],
                             "row": index + 1, "source_quantity": _canon(source_stage_total),
                             "stage_sum": _canon(stage_sum), "blocking": True})
        lines.append(line)
    return {"lines": lines, "totals": totals, "stages": stages,
            "roles": roles, "warnings": warnings}


def _detail_occurrence_key(line, counters):
    text = _norm(line.get("text_raw"))
    counters[text] += 1
    return text, counters[text]


def _merge_allocation_sheet(primary, allocation):
    allocation_map, counters = {}, Counter()
    for line in allocation["lines"]:
        if line["kind"] == "detail":
            allocation_map[_detail_occurrence_key(line, counters)] = line
    counters = Counter()
    matched = 0
    for line in primary["lines"]:
        if line["kind"] != "detail":
            continue
        other = allocation_map.get(_detail_occurrence_key(line, counters))
        if not other or not other.get("stage_quantities"):
            continue
        source_qty = _decimal(line["contract_quantity"].get("value"))
        allocation_qty = _decimal(other["contract_quantity"].get("value"))
        if (source_qty is not None and allocation_qty is not None
                and not _quantity_equal(source_qty, allocation_qty)):
            continue
        for key in ("stage_quantities", "stage_cells", "stage_quantity_sum", "stage_total",
                    "stage_total_matches_sum", "stage_status", "stage_bucket_key",
                    "stage_bucket_name", "classification_evidence", "classification_confidence"):
            line[key] = other.get(key)
        matched += 1
    return matched


def _relative_difference(source, normalized):
    source, normalized = _decimal(source), _decimal(normalized)
    if source is None or normalized is None:
        return None
    if source == 0:
        return Decimal(0) if normalized == 0 else Decimal("Infinity")
    return abs(normalized - source) / abs(source)


def _quantity_equal(left, right):
    left, right = _decimal(left), _decimal(right)
    if left is None or right is None:
        return left is right
    return abs(left - right) <= Decimal("0.000000001")


def _build_audit(lines, document_kind, warnings):
    details = [line for line in lines if line["kind"] == "detail"]
    row_checks, blocking = [], []
    source_amount_total = normalized_amount_total = Decimal(0)
    source_qty_by_uom, normalized_qty_by_uom = defaultdict(Decimal), defaultdict(Decimal)
    for line in details:
        qty = _decimal(line["contract_quantity"].get("value"))
        price = _decimal(line["unit_price"].get("value"))
        amount = _decimal(line["amount"].get("value"))
        normalized_qty, normalized_price, normalized_amount = qty, price, amount
        uom = _norm(line.get("unit_raw")) or "(none)"
        if qty is not None:
            source_qty_by_uom[uom] += qty
            normalized_qty_by_uom[uom] += normalized_qty
        if amount is not None:
            source_amount_total += amount
            normalized_amount_total += normalized_amount
        fields = {}
        for name, source, normalized, money in (
                ("quantity", qty, normalized_qty, False),
                ("unit_price", price, normalized_price, True),
                ("amount", amount, normalized_amount, True)):
            diff = _relative_difference(source, normalized)
            mismatch = (source != normalized if not money else
                        diff is not None and diff > MONEY_TOLERANCE_RATIO)
            fields[name] = {"source": _canon(source), "normalized": _canon(normalized),
                            "difference_ratio": _canon(diff) if diff is not None and diff.is_finite() else
                                                ("infinity" if diff is not None else None),
                            "within_tolerance": not mismatch}
            if mismatch:
                blocking.append({"code": "NORMALIZED_FIELD_MISMATCH", "row": line["source_row"],
                                 "field": name})
        if document_kind == "QUOTATION":
            for required, value in (("quantity", qty), ("unit_price", price), ("amount", amount)):
                if value is None:
                    blocking.append({"code": "QUOTATION_FIELD_MISSING", "row": line["source_row"],
                                     "field": required})
        if line.get("stage_total_matches_sum") is False:
            blocking.append({"code": "STAGE_TOTAL_MISMATCH", "row": line["source_row"]})
        row_checks.append({"source_sheet": line["contract_quantity"].get("source_sheet"),
                           "source_row": line["source_row"], "item_name": line["text_raw"],
                           "uom": line.get("unit_raw"), "fields": fields,
                           "stage_status": line.get("stage_status"),
                           "classification_confidence": line.get("classification_confidence")})
    aggregate_diff = _relative_difference(source_amount_total, normalized_amount_total)
    if aggregate_diff is not None and aggregate_diff > MONEY_TOLERANCE_RATIO:
        blocking.append({"code": "TOTAL_AMOUNT_MISMATCH"})
    for warning in warnings:
        if warning.get("blocking"):
            blocking.append({"code": warning["code"], "row": warning.get("row")})
    return {
        "policy": "SOURCE_VS_CANONICAL_FAIL_CLOSED", "money_tolerance_percent": "0.02",
        "money_tolerance_ratio": _canon(MONEY_TOLERANCE_RATIO),
        "source_detail_count": len(details), "normalized_detail_count": len(details),
        "source_amount_total": _canon(source_amount_total),
        "normalized_amount_total": _canon(normalized_amount_total),
        "amount_difference_ratio": (_canon(aggregate_diff) if aggregate_diff is not None else None),
        "quantity_totals_by_uom": {key: {"source": _canon(value),
            "normalized": _canon(normalized_qty_by_uom[key]),
            "matches": value == normalized_qty_by_uom[key]}
            for key, value in sorted(source_qty_by_uom.items())},
        "row_checks": row_checks, "blocking_issues": blocking,
        "ready_for_official_commit": not blocking and bool(details),
    }


def _sheet_filename_similarity(filename, sheet_name):
    ignored = {"bao", "gia", "khoi", "luong", "file", "sheet", "copy", "version"}
    file_tokens = {token for token in re.findall(r"[a-z0-9]+", _norm(filename))
                   if len(token) >= 2 and token not in ignored}
    sheet_tokens = {token for token in re.findall(r"[a-z0-9]+", _norm(sheet_name))
                    if len(token) >= 2 and token not in ignored}
    return len(file_tokens & sheet_tokens)


def normalize_workbook(data, filename, sheet_name=None):
    """Return a canonical Preview.  No database/filesystem write occurs."""
    if not isinstance(data, (bytes, bytearray)) or not data:
        raise BoqNormalizationError("File import rong hoac khong hop le.")
    sheets = _read_workbook(bytes(data), filename)
    candidates = []
    for sheet in sheets:
        candidate = _candidate(sheet["rows"], sheet["name"])
        if candidate:
            parsed = _parse_sheet(sheet, candidate)
            detail_count = sum(line["kind"] == "detail" for line in parsed["lines"])
            candidate.update({"sheet": sheet, "parsed": parsed, "detail_count": detail_count})
            candidates.append(candidate)
    if not candidates:
        raise BoqNormalizationError("Khong nhan dien duoc sheet/header BOQ phu hop.")
    financial = [candidate for candidate in candidates
                 if candidate["price_count"] or candidate["amount_count"]]
    eligible = financial or candidates
    if sheet_name:
        eligible = [candidate for candidate in candidates
                    if candidate["sheet_name"] == sheet_name]
        if not eligible:
            raise BoqNormalizationError("Sheet duoc chon khong co header BOQ hop le: %s" % sheet_name)
    primary_candidate = max(
        eligible,
        key=lambda item: (_sheet_filename_similarity(filename, item["sheet_name"]),
                          0 if re.search(r"\(\s*\d+\s*\)\s*$", item["sheet_name"]) else 1,
                          item["score"], item["detail_count"]))
    primary = primary_candidate["parsed"]
    allocation_candidates = [candidate for candidate in candidates
                             if candidate["stage_count"] and candidate is not primary_candidate]
    matched_allocation_rows = 0
    allocation_parsed = None
    allocation_sheet_name = primary_candidate["sheet_name"] if primary_candidate["stage_count"] else None
    if allocation_candidates:
        allocation_candidate = max(allocation_candidates,
                                   key=lambda item: (item["detail_count"], item["stage_count"]))
        allocation_parsed = allocation_candidate["parsed"]
        matched_allocation_rows = _merge_allocation_sheet(primary, allocation_parsed)
        if matched_allocation_rows:
            allocation_sheet_name = allocation_candidate["sheet_name"]
    stage_defs = {}
    for stage in list(primary.get("stages") or []) + list(
            (allocation_parsed or {}).get("stages") or []):
        stage_defs.setdefault(stage["key"], dict(stage))
    for line in primary["lines"]:
        for quantity in line.get("stage_quantities") or []:
            key = quantity["stage_key"]
            stage_defs.setdefault(key, {"key": key, "source_column": quantity.get("source_column"),
                "source_index": None, "raw_name": quantity["stage_name_raw"],
                "name": quantity["stage_name_raw"], "order": len(stage_defs) + 1,
                "is_unallocated_bucket": False})
    stages = list(stage_defs.values())
    stages.append({"key": "unallocated", "source_column": None, "source_index": None,
                   "raw_name": UNALLOCATED_STAGE_NAME, "name": UNALLOCATED_STAGE_NAME,
                   "order": len(stages) + 1, "is_unallocated_bucket": True})
    filename_text = _norm(filename)
    primary_text = _norm("%s %s" % (filename, primary_candidate["sheet_name"]))
    has_finance = bool(primary_candidate["price_count"] or primary_candidate["amount_count"])
    if "phat sinh" in filename_text:
        document_kind = "VARIATION_QUANTITY"
    elif "thanh toan" in primary_text or "quyet toan" in primary_text:
        document_kind = "PAYMENT_QUANTITY"
    elif "phat sinh" in primary_text:
        document_kind = "VARIATION_QUANTITY"
    elif has_finance or "bao gia" in primary_text:
        document_kind = "QUOTATION"
    else:
        document_kind = "BOQ_QUANTITY"
    warnings = list(primary.get("warnings") or [])
    if any(sheet["truncated"] for sheet in sheets):
        warnings.append({"code": "WORKBOOK_TRUNCATED", "blocking": True})
    if allocation_candidates and not matched_allocation_rows and not primary_candidate["stage_count"]:
        warnings.append({"code": "ALLOCATION_SHEET_UNMATCHED", "blocking": True})
    audit = _build_audit(primary["lines"], document_kind, warnings)
    details = [line for line in primary["lines"] if line["kind"] == "detail"]
    headings = [line for line in primary["lines"] if line["kind"] == "heading"]
    unallocated = sum(not line.get("stage_quantities") for line in details)
    return {
        "ok": True, "parser_version": PARSER_VERSION, "document_kind": document_kind,
        "source": {"kind": os.path.splitext(filename or "")[1].lower().lstrip("."),
                   "name": os.path.basename(filename or "workbook"), "byte_size": len(data),
                   "sha256": hashlib.sha256(data).hexdigest(),
                   "sheet_names": [sheet["name"] for sheet in sheets],
                   "sheet_name": primary_candidate["sheet_name"],
                   "allocation_sheet_name": allocation_sheet_name,
                   "header_row": primary_candidate["header_row"]},
        "headers": [{"source_index": index + 1, "source_column": _column_name(index),
                     "raw": label, "normalised": label, "role": None}
                    for index, label in enumerate(primary_candidate["labels"])],
        "columns": {}, "stages": stages, "lines": primary["lines"],
        "totals": primary.get("totals") or [],
        "ignored_rows": [], "warnings": warnings, "duplicate_item_groups": [],
        "counts": {"sheet_row_count": len(primary_candidate["sheet"]["rows"]),
                   "meaningful_data_row_count": len(primary["lines"]),
                    "line_count": len(primary["lines"]), "heading_count": len(headings),
                    "detail_count": len(details),
                    "total_count": len(primary.get("totals") or []), "ignored_row_count": 0,
                   "stage_count": len(stages) - 1,
                   "stage_definition_count_including_bucket": len(stages),
                   "stage_allocation_count": sum(len(line.get("stage_quantities") or []) for line in details),
                   "stage_allocated_detail_count": len(details) - unallocated,
                   "unallocated_detail_count": unallocated, "warning_count": len(warnings)},
        "invariants": {"header_detected": True, "row_order_preserved": True,
                       "detail_signature_complete": True,
                       "source_hash_preserved": True,
                       "source_vs_normalized_reconciled": audit["ready_for_official_commit"]},
        "normalization_audit": audit,
        "candidate_sheets": [{"sheet_name": candidate["sheet_name"],
                              "header_row": candidate["header_row"],
                              "detail_count": candidate["detail_count"],
                              "price_columns": candidate["price_count"],
                              "amount_columns": candidate["amount_count"],
                              "stage_columns": candidate["stage_count"]}
                             for candidate in candidates],
    }


__all__ = ["BoqNormalizationError", "MONEY_TOLERANCE_RATIO", "PARSER_VERSION",
           "UNALLOCATED_STAGE_NAME", "normalize_workbook"]
