# -*- coding: utf-8 -*-
"""WO-10 — Import 3 file Excel lam NGUON CHUAN + doi chieu bao gia <-> hoa don.

- import_customers(path, commit): danh ba khach (MST chong trung, 2 pha preview/commit)
- import_invoices(path, commit):  hoa don ban ra (gom theo MaHD, chong trung, tu tao khach)
- doi_chieu(conn):                bao gia co hoa don khop = XONG
Chay CLI:  python import_excel.py customers|invoices|doichieu [duong_dan] [--commit]
"""
import os
import re
import sys
import unicodedata
from datetime import datetime

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

import db as D
from api_write import ValidationError

BASE = r"D:\Quản trị DOANH NGHIỆP"
DEFAULT_CUSTOMER_FILE = os.path.join(BASE, "Customer data.xlsx")
DEFAULT_INVOICE_FILES = [os.path.join(BASE, "Invoice_20260708.xlsx"),
                         os.path.join(BASE, "InvoiceTTH_20260708.xlsx")]

# FIND-007 (P1): duong dan Excel truyen tu client (path/paths trong /api/import_run)
# khong duoc kiem soat truoc khi toi openpyxl.load_workbook(). Gioi han dung luong
# o day mirror UPLOAD_MAX (api_write.py, 15MB) va cac gioi han 8MB/12MB o cac luong
# import khac -- 20MB vi bang tinh nghiep vu (danh ba/hoa don) co the lon hon 1 sao ke.
IMPORT_MAX_BYTES = 20 * 1024 * 1024


# ---------- helpers ----------
def clean_str(v):
    """Bo lone-surrogate/ky tu hong tu file Excel nguon (tranh vo JSON/console)."""
    if v is None:
        return ""
    return str(v).encode("utf-8", "ignore").decode("utf-8", "ignore").strip()


def norm_mst(v):
    """MST chuan hoa = chi giu chu so (va dau - cho MST chi nhanh 13 so)."""
    if v is None:
        return ""
    return re.sub(r"[^0-9]", "", str(v))


def norm_name(s):
    if not s:
        return ""
    r = "".join(c for c in unicodedata.normalize("NFD", str(s))
                if unicodedata.category(c) != "Mn")
    r = r.replace("đ", "d").replace("Đ", "D").lower()
    r = re.sub(r"\b(cong ty|cty|cong|tnhh|co phan|cp|mtv|hưu han|huu han|company|co\.|ltd)\b", " ", r)
    r = re.sub(r"\b(20\d\d)\b", " ", r)
    return re.sub(r"\s+", " ", r).strip()


def parse_date_ddmmyyyy(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    s = str(v).strip()
    m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", s)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            return datetime(y, mo, d).date().isoformat()
        except ValueError:
            return None
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        return s[:10]
    return None


def to_num(v):
    if v is None or v == "":
        return 0.0
    try:
        return float(str(v).replace(",", ""))
    except ValueError:
        return 0.0


def _find_file(path):
    """Chiu duoc ten file co khoang trang thua o cuoi.

    FIND-007 (P1): day la NOI DUY NHAT ma path/paths client-supplied cua
    /api/import_run di qua truoc khi toi openpyxl.load_workbook(). Chan UNC,
    bat buoc duoi .xlsx, va gioi han dung luong O DAY -- truoc khi tra ve
    duong dan cho ham goi (import_customers/import_invoices) mo file.
    """
    if re.match(r"^[/\\]{2,}", path):
        raise ValidationError("Khong chap nhan duong dan mang (UNC): %s" % path)
    if not path.lower().endswith(".xlsx"):
        raise ValidationError("Chi chap nhan file .xlsx: %s" % path)
    if os.path.isfile(path):
        found = path
    else:
        found = None
        d, base = os.path.dirname(path), os.path.basename(path).strip().lower()
        if os.path.isdir(d):
            for f in os.listdir(d):
                if f.strip().lower() == base:
                    found = os.path.join(d, f)
                    break
        if found is None:
            raise FileNotFoundError(path)
    if os.path.getsize(found) > IMPORT_MAX_BYTES:
        raise ValidationError("File qua lon (>%dMB): %s"
                              % (IMPORT_MAX_BYTES // (1024 * 1024), found))
    return found


# ---------- 1.1 Import danh ba khach ----------
def import_customers(path=None, commit=False, conn=None):
    import openpyxl
    path = _find_file(path or DEFAULT_CUSTOMER_FILE)
    own = conn is None
    conn = conn or D.get_conn()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    header, data = rows[0], rows[1:]

    # index khach hien co theo MST + ten chuan hoa
    by_mst, by_name = {}, {}
    for r in conn.execute("SELECT id, customer_name, tax_id FROM customer"):
        m = norm_mst(r["tax_id"])
        if m:
            by_mst[m] = r["id"]
        by_name.setdefault(norm_name(r["customer_name"]), r["id"])

    stats = {"them_moi": 0, "cap_nhat": 0, "trung_giu_nguyen": 0, "can_soat": 0, "dong_rong": 0}
    details = []
    seq = conn.execute("SELECT COUNT(*) FROM customer").fetchone()[0]

    for row in data:
        name = (row[1] or "").strip() if len(row) > 1 and row[1] else ""
        if not name:
            stats["dong_rong"] += 1
            continue
        contact = (row[2] or "").strip() if len(row) > 2 and row[2] else ""
        mst = norm_mst(row[3]) if len(row) > 3 else ""
        mail = (row[4] or "").strip() if len(row) > 4 and row[4] else ""
        sdt = str(row[5]).strip() if len(row) > 5 and row[5] else ""
        notes = (row[6] or "").strip() if len(row) > 6 and row[6] else ""

        cid = by_mst.get(mst) if mst else None
        if cid is None and not mst:
            cid = by_name.get(norm_name(name))
        if cid:
            # cap nhat thong tin thieu, KHONG ghi de ten
            if commit:
                cur = conn.execute("SELECT * FROM customer WHERE id=?", (cid,)).fetchone()
                sets, vals = [], []
                for col, val in [("nguoi_lien_he", contact), ("email", mail),
                                 ("dien_thoai", sdt), ("ghi_chu", notes), ("tax_id", mst or None)]:
                    if val and not cur[col]:
                        sets.append("%s=?" % col)
                        vals.append(val)
                if sets:
                    vals.append(cid)
                    conn.execute("UPDATE customer SET %s WHERE id=?" % ",".join(sets), vals)
                    stats["cap_nhat"] += 1
                else:
                    stats["trung_giu_nguyen"] += 1
            else:
                stats["cap_nhat"] += 1
            details.append(("cap_nhat", name, mst))
        else:
            stats["them_moi"] += 1
            details.append(("them_moi", name, mst))
            if commit:
                seq += 1
                conn.execute(
                    """INSERT INTO customer(code, customer_name, tax_id, nguoi_lien_he, email,
                       dien_thoai, ghi_chu, nguon) VALUES(?,?,?,?,?,?,?,?)""",
                    ("KH-M-%04d" % seq, name, mst or None, contact, mail, sdt, notes, "master_xlsx"))
                new_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
                if mst:
                    by_mst[mst] = new_id
                by_name[norm_name(name)] = new_id
    if commit:
        conn.commit()
    total = conn.execute("SELECT COUNT(*) FROM customer").fetchone()[0]
    out = {"file": path, "stats": stats, "tong_khach_sau": total, "commit": commit,
           "preview": details[:30]}
    if own:
        conn.close()
    return out


# ---------- 1.2 Import hoa don ban ra ----------
def import_invoices(paths=None, commit=False, chieu="ban_ra", conn=None):
    import openpyxl
    paths = paths or DEFAULT_INVOICE_FILES
    if isinstance(paths, str):
        paths = [paths]
    own = conn is None
    conn = conn or D.get_conn()

    by_mst = {}
    for r in conn.execute("SELECT id, tax_id FROM customer WHERE tax_id IS NOT NULL"):
        m = norm_mst(r["tax_id"])
        if m:
            by_mst[m] = r["id"]

    stats = {"hoa_don_moi": 0, "trung_bo_qua": 0, "dong_hang": 0, "khach_tu_hd": 0,
             "tong_tien": 0.0}
    for path in paths:
        path = _find_file(path)
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        header = [str(h or "").strip() for h in rows[0]]
        idx = {h: i for i, h in enumerate(header)}

        def g(row, col):
            i = idx.get(col)
            return row[i] if i is not None and i < len(row) else None

        # gom theo MaHD (trong pham vi 1 file)
        groups = {}
        for row in rows[1:]:
            ma = str(g(row, "MaHD") or "").strip()
            if not ma:
                continue
            groups.setdefault(ma, []).append(row)

        fname = os.path.basename(path)
        for ma, lines in groups.items():
            r0 = lines[0]
            ngay = parse_date_ddmmyyyy(g(r0, "NgayHoaDon"))
            mst = norm_mst(g(r0, "MaSoThue"))
            ten = str(g(r0, "TenDonVi") or "").strip()
            dia_chi = str(g(r0, "DiaChiKhachHang") or "").strip()
            httt = str(g(r0, "HinhThucThanhToan") or "").strip()
            tong_tt = sum(to_num(g(r, "ThanhTien")) for r in lines)
            tong_thue = sum(to_num(g(r, "TienThue")) for r in lines)

            # chong trung (ma_hd + ngay + mst) — ma_hd chi unique trong 1 dot xuat
            dup = conn.execute("SELECT id FROM hoa_don WHERE ma_hd=? AND ngay=? AND mst=?",
                               (ma, ngay, mst)).fetchone()
            if dup:
                stats["trung_bo_qua"] += 1
                continue

            # khach: khop MST, chua co -> tu tao tu hoa don
            cid = by_mst.get(mst)
            if not cid and commit:
                n = conn.execute("SELECT COUNT(*) FROM customer").fetchone()[0] + 1
                conn.execute(
                    """INSERT INTO customer(code, customer_name, tax_id, dia_chi, nguon)
                       VALUES(?,?,?,?,?)""",
                    ("KH-HD-%04d" % n, ten or ("MST " + mst), mst or None, dia_chi, "tu_hoa_don"))
                cid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
                if mst:
                    by_mst[mst] = cid
                stats["khach_tu_hd"] += 1
            elif not cid:
                stats["khach_tu_hd"] += 1  # dem preview

            stats["hoa_don_moi"] += 1
            stats["tong_tien"] += tong_tt + tong_thue
            if not commit:
                continue
            conn.execute(
                """INSERT INTO hoa_don(ma_hd, ngay, customer_id, mst, ten_don_vi, dia_chi,
                   tong_truoc_thue, tong_thue, tong_cong, hinh_thuc_tt, chieu, nguon_file)
                   VALUES(?,?,?,?,?,?,?,?,?,?,?,?)""",
                (ma, ngay, cid, mst, ten, dia_chi, tong_tt, tong_thue,
                 tong_tt + tong_thue, httt, chieu, fname))
            hd_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
            for r in lines:
                ten_hang = str(g(r, "TenHangHoa") or "").strip()
                conn.execute(
                    """INSERT INTO hoa_don_dong(hoa_don_id, so_tt, ma_hang, ten_hang_hoa, dvt,
                       so_luong, don_gia, thanh_tien, thue_suat, tien_thue)
                       VALUES(?,?,?,?,?,?,?,?,?,?)""",
                    (hd_id, g(r, "SoTT"), g(r, "MaHang"), ten_hang, g(r, "DVT"),
                     to_num(g(r, "SoLuong")), to_num(g(r, "DonGia")), to_num(g(r, "ThanhTien")),
                     str(g(r, "ThueSuat") or ""), to_num(g(r, "TienThue"))))
                stats["dong_hang"] += 1
                # lam giau catalog mat hang
                if ten_hang:
                    conn.execute(
                        """INSERT INTO mat_hang_tu_hoa_don(ten_hang_hoa, dvt, gia_gan_nhat, lan_gan_nhat)
                           VALUES(?,?,?,?)
                           ON CONFLICT(ten_hang_hoa) DO UPDATE SET
                             so_lan_ban = so_lan_ban + 1,
                             gia_gan_nhat = excluded.gia_gan_nhat,
                             lan_gan_nhat = excluded.lan_gan_nhat""",
                        (ten_hang, str(g(r, "DVT") or ""), to_num(g(r, "DonGia")), ngay))
    if commit:
        conn.commit()
    out = {"files": [os.path.basename(_find_file(p)) for p in paths], "stats": stats, "commit": commit}
    if own:
        conn.close()
    return out


# ---------- 2.1 Doi chieu bao gia <-> hoa don ----------
TOLERANCE = 0.15  # ±15% (bao gia thuong chua VAT / phat sinh nho)


def doi_chieu(conn=None):
    """Bao gia nao co hoa don ban ra cung khach (MST) tien ~khop, ngay HD >= ngay bao gia
    -> danh dau XONG. Khop mo -> can_xac_nhan."""
    own = conn is None
    conn = conn or D.get_conn()
    stats = {"xong": 0, "can_xac_nhan": 0, "chua": 0}
    quotes = conn.execute("""SELECT q.id, q.code, q.grand_total, q.ngay_lap, q.customer_id,
                             c.tax_id FROM quotation q
                             LEFT JOIN customer c ON c.id=q.customer_id""").fetchall()
    for q in quotes:
        mst = norm_mst(q["tax_id"])
        matched, fuzzy = None, 0
        if mst and q["grand_total"]:
            cands = conn.execute(
                """SELECT id, tong_cong, tong_truoc_thue, ngay FROM hoa_don
                   WHERE mst=? AND chieu='ban_ra' AND (ngay >= ? OR ? IS NULL)""",
                (mst, q["ngay_lap"], q["ngay_lap"])).fetchall()
            for hd in cands:
                for tien in (hd["tong_cong"], hd["tong_truoc_thue"]):
                    if tien and abs(tien - q["grand_total"]) / q["grand_total"] <= TOLERANCE:
                        fuzzy += 1
                        matched = matched or hd["id"]
                        break
        if matched and fuzzy == 1:
            tt = "xong"
        elif matched:
            tt = "can_xac_nhan"   # nhieu ung vien — nguoi duyet gan tay
        else:
            tt = "chua"
        stats[tt] += 1
        conn.execute("UPDATE quotation SET hoa_don_lien_ket=?, trang_thai_doi_chieu=? WHERE id=?",
                     (matched if tt == "xong" else None, tt, q["id"]))
    conn.commit()
    if own:
        conn.close()
    return stats


if __name__ == "__main__":
    cmd = sys.argv[1] if len(sys.argv) > 1 else "customers"
    commit = "--commit" in sys.argv
    args = [a for a in sys.argv[2:] if a != "--commit"]
    if cmd == "customers":
        r = import_customers(args[0] if args else None, commit=commit)
    elif cmd == "invoices":
        r = import_invoices(args if args else None, commit=commit)
    elif cmd == "doichieu":
        r = doi_chieu()
    else:
        print("customers|invoices|doichieu [path] [--commit]")
        sys.exit(1)
    import json
    print(json.dumps(r, ensure_ascii=False, indent=1, default=str))
