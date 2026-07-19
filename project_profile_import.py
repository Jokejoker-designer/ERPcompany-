# -*- coding: utf-8 -*-
"""Pure XLSX parser for construction project-profile previews.

This module deliberately has no database, HTTP, server, or application-state
dependencies.  It accepts XLSX bytes (or a local path), reads the workbook in
formula and cached-value modes, and returns a JSON-serializable preview.  A
later, explicitly authorised layer may decide how to persist that preview.

The source workbook remains authoritative: raw text, row order, formulas,
stage columns, contract quantity, and stage total are kept separately.  The
parser never invents a floor allocation for a line that has none.
"""

from __future__ import annotations

import ast
import hashlib
import io
import os
import re
import unicodedata
from collections import defaultdict
from decimal import Decimal, InvalidOperation, localcontext
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Optional, Sequence, Tuple, Union

import openpyxl
from openpyxl.utils import get_column_letter, range_boundaries


UNALLOCATED_STAGE_NAME = "Chưa phân tầng/Toàn công trình"
PARSER_VERSION = "1.0"

_NULL_MARKERS = {"-", "–", "—"}
_CELL_REF_RE = re.compile(r"^[A-Z]{1,3}[1-9][0-9]*$")
_INTEGER_TOKEN_RE = re.compile(r"^[0-9]+$")
_DECIMAL_TOKEN_RE = re.compile(r"^([0-9]+)(?:\.[0-9]+)+$")
_ROMAN_TOKENS = {
    "I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X", "XI", "XII"
}


class ProjectProfileImportError(ValueError):
    """Raised when an XLSX cannot be interpreted as a project-profile quote."""


Source = Union[bytes, bytearray, memoryview, str, os.PathLike]


def _normalise_key(value: Any) -> str:
    """Accent/case/punctuation-insensitive key used only for detection."""
    if value is None:
        return ""
    text = str(value).replace("đ", "d").replace("Đ", "D")
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.casefold()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(text.split())


def _canonical_decimal(value: Decimal) -> str:
    """Return a non-exponent JSON-safe decimal string."""
    if value == 0:
        return "0"
    rendered = format(value, "f")
    if "." in rendered:
        rendered = rendered.rstrip("0").rstrip(".")
    return rendered


def _as_decimal(value: Any, *, allow_text: bool = True) -> Optional[Decimal]:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, int):
        return Decimal(value)
    if isinstance(value, float):
        # str(float) avoids importing the binary representation into Decimal.
        return Decimal(str(value))
    if allow_text and isinstance(value, str):
        text = value.strip()
        if not text or text in _NULL_MARKERS:
            return None
        # XLSX numeric text in this workflow uses dot decimals.  Thousands
        # separators are intentionally not guessed: ambiguity must surface.
        if not re.fullmatch(r"[+-]?(?:[0-9]+(?:\.[0-9]*)?|\.[0-9]+)", text):
            return None
        try:
            return Decimal(text)
        except InvalidOperation:
            return None
    return None


def _raw_json_value(value: Any) -> Any:
    """Keep strings byte-for-byte while making numerics canonical strings."""
    number = _as_decimal(value, allow_text=False)
    if number is not None:
        return _canonical_decimal(number)
    if value is None or isinstance(value, (str, bool)):
        return value
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value)


def _read_source(source: Source) -> Tuple[bytes, Optional[str], str]:
    if isinstance(source, (bytes, bytearray, memoryview)):
        data = bytes(source)
        source_name = None
        source_kind = "bytes"
    elif isinstance(source, (str, os.PathLike)):
        path = Path(os.fspath(source))
        try:
            data = path.read_bytes()
        except OSError as exc:
            raise ProjectProfileImportError("Không đọc được file XLSX: %s" % exc) from exc
        source_name = path.name
        source_kind = "path"
    else:
        raise TypeError("source phải là XLSX bytes hoặc path")
    if not data:
        raise ProjectProfileImportError("File XLSX rỗng.")
    return data, source_name, source_kind


def _header_role_map(ws: Any, row: int) -> Optional[Dict[str, int]]:
    values = {column: _normalise_key(ws.cell(row, column).value)
              for column in range(1, ws.max_column + 1)}

    def first(predicate: Any) -> Optional[int]:
        for column, key in values.items():
            if key and predicate(key):
                return column
        return None

    roles: Dict[str, Optional[int]] = {
        "stt": first(lambda key: key == "stt"),
        "description": first(lambda key: "hang muc cong viec" in key or key == "hang muc"),
        "technical": first(lambda key: "chi dan ky thuat" in key or
                            ("yeu cau" in key and "ky thuat" in key)),
        "stage_total": first(lambda key: "tong" in key and "khoi luong" in key),
        "unit": first(lambda key: "don vi tinh" in key or key in {"dvt", "don vi"}),
        "contract_quantity": first(lambda key: key in {"khoi luong", "so luong"}),
        "unit_price": first(lambda key: "don gia" in key),
        "amount": first(lambda key: "thanh tien" in key),
        "category": first(lambda key: "chung loai" in key or "nhan hieu" in key),
        "note": first(lambda key: "ghi chu" in key),
    }
    required = (
        "stt", "description", "technical", "stage_total", "unit",
        "contract_quantity", "unit_price", "amount",
    )
    if any(roles[name] is None for name in required):
        return None
    result = {name: int(column) for name, column in roles.items() if column is not None}
    if not (
        result["description"] < result["technical"] < result["stage_total"]
        < result["unit"] < result["contract_quantity"]
        < result["unit_price"] < result["amount"]
    ):
        return None
    return result


def _detect_header(ws: Any, max_scan_rows: int = 50) -> Tuple[int, Dict[str, int]]:
    for row in range(1, min(ws.max_row, max_scan_rows) + 1):
        roles = _header_role_map(ws, row)
        if roles is not None:
            return row, roles
    raise ProjectProfileImportError(
        "Không tìm thấy header báo giá có STT/hạng mục/chỉ dẫn/tổng KL/đơn vị/KL/đơn giá/thành tiền."
    )


class _FormulaResolver:
    """Small, non-executing evaluator for arithmetic used by quote sheets."""

    def __init__(self, formula_ws: Any, cached_ws: Any):
        self.formula_ws = formula_ws
        self.cached_ws = cached_ws
        self._memo: Dict[str, Optional[Decimal]] = {}
        self.unsupported: Dict[str, str] = {}

    def cached_decimal(self, coordinate: str) -> Optional[Decimal]:
        return _as_decimal(self.cached_ws[coordinate].value)

    def decimal(self, coordinate: str, stack: Optional[set] = None) -> Optional[Decimal]:
        coordinate = coordinate.upper()
        if coordinate in self._memo:
            return self._memo[coordinate]
        if stack is None:
            stack = set()
        if coordinate in stack:
            self.unsupported[coordinate] = "circular formula"
            return self.cached_decimal(coordinate)

        cell = self.formula_ws[coordinate]
        if cell.data_type != "f":
            value = _as_decimal(cell.value)
            self._memo[coordinate] = value
            return value

        formula = str(cell.value or "")
        try:
            computed = self._evaluate(formula, stack | {coordinate})
        except (InvalidOperation, ValueError, ZeroDivisionError, SyntaxError) as exc:
            self.unsupported[coordinate] = str(exc)
            computed = None
        if computed is None:
            computed = self.cached_decimal(coordinate)
        self._memo[coordinate] = computed
        return computed

    def _evaluate(self, formula: str, stack: set) -> Optional[Decimal]:
        expression = formula.strip()
        if expression.startswith("="):
            expression = expression[1:].strip()
        if not expression:
            return None

        sum_match = re.fullmatch(r"SUM\((.*)\)", expression, flags=re.IGNORECASE)
        if sum_match:
            inside = sum_match.group(1).strip()
            if not inside:
                return Decimal(0)
            total = Decimal(0)
            for argument in [part.strip() for part in inside.split(",")]:
                range_match = re.fullmatch(
                    r"([A-Z]{1,3}[1-9][0-9]*):([A-Z]{1,3}[1-9][0-9]*)",
                    argument,
                    flags=re.IGNORECASE,
                )
                if range_match:
                    min_col, min_row, max_col, max_row = range_boundaries(
                        "%s:%s" % (range_match.group(1), range_match.group(2))
                    )
                    for row in range(min_row, max_row + 1):
                        for column in range(min_col, max_col + 1):
                            value = self.decimal("%s%s" % (get_column_letter(column), row), stack)
                            if value is not None:
                                total += value
                elif _CELL_REF_RE.fullmatch(argument.upper()):
                    value = self.decimal(argument.upper(), stack)
                    if value is not None:
                        total += value
                else:
                    value = self._evaluate_ast(argument, stack)
                    if value is not None:
                        total += value
            return total
        return self._evaluate_ast(expression, stack)

    def _evaluate_ast(self, expression: str, stack: set) -> Optional[Decimal]:
        # Excel's 10% is arithmetic percentage, not Python's modulo operator.
        expression = re.sub(
            r"(?<![A-Za-z0-9_.])([0-9]+(?:\.[0-9]+)?)\s*%",
            r"(\1/100)",
            expression,
        )
        tree = ast.parse(expression, mode="eval")

        def visit(node: ast.AST) -> Decimal:
            if isinstance(node, ast.Expression):
                return visit(node.body)
            if isinstance(node, ast.Constant) and isinstance(node.value, (int, float)) \
                    and not isinstance(node.value, bool):
                return Decimal(str(node.value))
            if isinstance(node, ast.Name) and _CELL_REF_RE.fullmatch(node.id.upper()):
                value = self.decimal(node.id.upper(), stack)
                if value is None:
                    return Decimal(0)
                return value
            if isinstance(node, ast.UnaryOp) and isinstance(node.op, (ast.UAdd, ast.USub)):
                value = visit(node.operand)
                return value if isinstance(node.op, ast.UAdd) else -value
            if isinstance(node, ast.BinOp) and isinstance(
                node.op, (ast.Add, ast.Sub, ast.Mult, ast.Div)
            ):
                left, right = visit(node.left), visit(node.right)
                if isinstance(node.op, ast.Add):
                    return left + right
                if isinstance(node.op, ast.Sub):
                    return left - right
                if isinstance(node.op, ast.Mult):
                    return left * right
                return left / right
            raise ValueError("unsupported formula expression")

        with localcontext() as context:
            context.prec = 40
            return visit(tree)


def _cell_payload(
    formula_ws: Any,
    cached_ws: Any,
    resolver: _FormulaResolver,
    row: int,
    column: int,
    *,
    coerce_numeric_text: bool = False,
    null_markers: bool = False,
) -> Dict[str, Any]:
    formula_cell = formula_ws.cell(row, column)
    cached_cell = cached_ws.cell(row, column)
    raw = formula_cell.value
    coordinate = formula_cell.coordinate
    formula = str(raw) if formula_cell.data_type == "f" else None
    cached = resolver.cached_decimal(coordinate) if formula else None
    coerced = False
    null_marker = False

    if isinstance(raw, str) and formula is None:
        stripped = raw.strip()
        if null_markers and stripped in _NULL_MARKERS:
            null_marker = True
            value = None
            source = "null_marker"
        elif coerce_numeric_text and _as_decimal(raw) is not None:
            value = _as_decimal(raw)
            coerced = True
            source = "text_coerced"
        else:
            value = _as_decimal(raw, allow_text=False)
            source = "literal" if value is not None else "non_numeric"
    elif formula:
        value = resolver.decimal(coordinate)
        source = "formula_recomputed" if coordinate not in resolver.unsupported else "formula_cached"
    else:
        value = _as_decimal(raw, allow_text=False)
        source = "literal" if value is not None else "missing"

    return {
        "source_cell": coordinate,
        "raw_value": _raw_json_value(raw),
        "formula": formula,
        "cached_value": _canonical_decimal(cached) if cached is not None else None,
        "value": _canonical_decimal(value) if value is not None else None,
        "value_source": source,
        "number_format": formula_cell.number_format,
        "coerced_from_text": coerced,
        "null_marker": null_marker,
    }


def _has_value(formula_ws: Any, cached_ws: Any, row: int, column: int) -> bool:
    for cell in (formula_ws.cell(row, column), cached_ws.cell(row, column)):
        value = cell.value
        if value is not None and (not isinstance(value, str) or value.strip()):
            return True
    return False


def _detail_signature(formula_ws: Any, cached_ws: Any, row: int, roles: Mapping[str, int]) -> bool:
    # Exact semantic equivalent of B/M/N/O/P in the official workbook.
    return all(
        _has_value(formula_ws, cached_ws, row, roles[name])
        for name in ("description", "unit", "contract_quantity", "unit_price", "amount")
    )


def _is_total_label(stt: Any, description: Any) -> bool:
    stt_key = _normalise_key(stt)
    description_key = _normalise_key(description)
    if re.fullmatch(r"tc[0-9]+", stt_key.replace(" ", "")):
        return True
    return description_key.startswith(("tong cong", "thue vat", "thanh tien"))


def _heading_level(stt: Any, stack: Sequence[Mapping[str, Any]]) -> int:
    token = str(stt or "").strip().upper()
    if token and set(token) == {"*"}:
        return 1
    if token in _ROMAN_TOKENS:
        alphabetic = [entry for entry in stack if entry.get("token_kind") == "alpha"]
        return int(alphabetic[-1]["level"]) + 1 if alphabetic else 2
    if re.fullmatch(r"[A-Z]", token):
        return 2
    if _INTEGER_TOKEN_RE.fullmatch(token):
        romans = [entry for entry in stack if entry.get("token_kind") == "roman"]
        if romans:
            return int(romans[-1]["level"]) + 1
        return (int(stack[-1]["level"]) + 1) if stack else 1
    if _DECIMAL_TOKEN_RE.fullmatch(token):
        return (int(stack[-1]["level"]) + 1) if stack else 1
    return int(stack[-1]["level"]) if stack else 1


def _token_kind(stt: Any) -> str:
    token = str(stt or "").strip().upper()
    if token and set(token) == {"*"}:
        return "marker"
    if token in _ROMAN_TOKENS:
        return "roman"
    if re.fullmatch(r"[A-Z]", token):
        return "alpha"
    if _INTEGER_TOKEN_RE.fullmatch(token):
        return "integer"
    if _DECIMAL_TOKEN_RE.fullmatch(token):
        return "decimal"
    return "other"


def _detail_hierarchy(stack: Sequence[Mapping[str, Any]], stt: Any) -> List[Dict[str, Any]]:
    result = [dict(entry) for entry in stack]
    token = str(stt or "").strip()
    if result and result[-1].get("token_kind") == "integer":
        parent_token = str(result[-1].get("stt_raw") or "").strip()
        decimal_match = _DECIMAL_TOKEN_RE.fullmatch(token)
        # A detail numbered 2.1 belongs below heading 2.  A detail numbered
        # simply 3 is a sibling of heading 2 and must not inherit it.
        if not decimal_match or decimal_match.group(1) != parent_token:
            result.pop()
    return result


def _public_hierarchy(entries: Sequence[Mapping[str, Any]]) -> List[Dict[str, Any]]:
    return [
        {
            "source_row": entry["source_row"],
            "level": entry["level"],
            "stt_raw": entry["stt_raw"],
            "text_raw": entry["text_raw"],
        }
        for entry in entries
    ]


def preview_project_profile_xlsx(source: Source, *, sheet_name: Optional[str] = None) -> Dict[str, Any]:
    """Parse an XLSX into a pure, JSON-serializable project-profile preview.

    No persistence occurs.  ``source`` may be bytes-like or a local path.
    Formula text and cached values are loaded from two independent workbook
    instances.  If ``sheet_name`` is omitted, the first sheet with a recognised
    quote header is selected.
    """
    data, source_name, source_kind = _read_source(source)
    digest = hashlib.sha256(data).hexdigest()

    try:
        formula_wb = openpyxl.load_workbook(
            io.BytesIO(data), data_only=False, read_only=False, keep_links=False
        )
        cached_wb = openpyxl.load_workbook(
            io.BytesIO(data), data_only=True, read_only=False, keep_links=False
        )
    except Exception as exc:
        raise ProjectProfileImportError("XLSX không hợp lệ: %s" % exc) from exc

    try:
        selected: Optional[Tuple[str, int, Dict[str, int]]] = None
        candidate_names = [sheet_name] if sheet_name else list(formula_wb.sheetnames)
        for candidate in candidate_names:
            if candidate not in formula_wb.sheetnames:
                continue
            try:
                header_row, roles = _detect_header(formula_wb[candidate])
            except ProjectProfileImportError:
                continue
            selected = (candidate, header_row, roles)
            break
        if selected is None:
            if sheet_name and sheet_name not in formula_wb.sheetnames:
                raise ProjectProfileImportError("Không có sheet %r." % sheet_name)
            raise ProjectProfileImportError("Không sheet nào có header báo giá hợp lệ.")

        selected_sheet, header_row, roles = selected
        formula_ws = formula_wb[selected_sheet]
        cached_ws = cached_wb[selected_sheet]
        resolver = _FormulaResolver(formula_ws, cached_ws)
        warnings: List[Dict[str, Any]] = []

        max_column = formula_ws.max_column
        headers: List[Dict[str, Any]] = []
        reverse_roles = {column: role for role, column in roles.items()}
        for column in range(1, max_column + 1):
            headers.append({
                "source_column": get_column_letter(column),
                "source_index": column,
                "raw": _raw_json_value(formula_ws.cell(header_row, column).value),
                "normalised": _normalise_key(formula_ws.cell(header_row, column).value),
                "role": reverse_roles.get(column),
            })

        stage_indexes = list(range(roles["technical"] + 1, roles["stage_total"]))
        if not stage_indexes:
            raise ProjectProfileImportError("Không có cột giai đoạn/tầng giữa chỉ dẫn và tổng KL.")
        stages: List[Dict[str, Any]] = []
        for order, column in enumerate(stage_indexes, start=1):
            raw_name = _raw_json_value(formula_ws.cell(header_row, column).value)
            if raw_name is None or not str(raw_name).strip():
                warnings.append({
                    "code": "EMPTY_STAGE_HEADER",
                    "row": header_row,
                    "cell": "%s%s" % (get_column_letter(column), header_row),
                    "message": "Cột giai đoạn không có tên; vẫn giữ theo vị trí nguồn.",
                })
            stages.append({
                "key": "stage:%s" % get_column_letter(column),
                "source_column": get_column_letter(column),
                "source_index": column,
                "raw_name": raw_name,
                "name": str(raw_name) if raw_name is not None else "Cột %s" % get_column_letter(column),
                "order": order,
                "is_unallocated_bucket": False,
            })
        stages.append({
            "key": "unallocated",
            "source_column": None,
            "source_index": None,
            "raw_name": UNALLOCATED_STAGE_NAME,
            "name": UNALLOCATED_STAGE_NAME,
            "order": len(stages) + 1,
            "is_unallocated_bucket": True,
        })

        formula_cells = [
            cell for row in formula_ws.iter_rows() for cell in row if cell.data_type == "f"
        ]
        missing_formula_cache = []
        for cell in formula_cells:
            if cached_ws[cell.coordinate].value is None:
                missing_formula_cache.append(cell.coordinate)
                warnings.append({
                    "code": "FORMULA_CACHE_MISSING",
                    "row": cell.row,
                    "cell": cell.coordinate,
                    "message": "Công thức không có cached value; parser chỉ tính lại phép toán an toàn được hỗ trợ.",
                })

        lines: List[Dict[str, Any]] = []
        totals: List[Dict[str, Any]] = []
        hierarchy_stack: List[Dict[str, Any]] = []
        ignored_rows: List[int] = []
        meaningful_rows = 0
        detail_rows: List[int] = []
        heading_rows: List[int] = []
        numeric_stage_cell_count = 0
        nonblank_stage_cell_count = 0
        stage_allocated_detail_count = 0
        unallocated_detail_count = 0

        for row in range(header_row + 1, formula_ws.max_row + 1):
            row_has_value = any(
                _has_value(formula_ws, cached_ws, row, column)
                for column in range(1, max_column + 1)
            )
            if not row_has_value:
                continue
            meaningful_rows += 1
            stt_value = formula_ws.cell(row, roles["stt"]).value
            description_value = formula_ws.cell(row, roles["description"]).value
            stt_raw = _raw_json_value(stt_value)
            text_raw = _raw_json_value(description_value)

            is_detail = _detail_signature(formula_ws, cached_ws, row, roles)
            if is_detail:
                detail_rows.append(row)
                parents = _detail_hierarchy(hierarchy_stack, stt_raw)
                stage_cells: List[Dict[str, Any]] = []
                stage_quantities: List[Dict[str, Any]] = []
                stage_sum = Decimal(0)
                for stage in stages[:-1]:
                    column = int(stage["source_index"])
                    if not _has_value(formula_ws, cached_ws, row, column):
                        continue
                    nonblank_stage_cell_count += 1
                    payload = _cell_payload(
                        formula_ws,
                        cached_ws,
                        resolver,
                        row,
                        column,
                        coerce_numeric_text=True,
                        null_markers=True,
                    )
                    stage_cell = {
                        "stage_key": stage["key"],
                        "stage_name_raw": stage["raw_name"],
                        # Convenience alias for consumers rendering the
                        # stage matrix; ``value`` remains in the full cell
                        # payload for a uniform formula/cached-value schema.
                        "quantity": payload["value"],
                        **payload,
                    }
                    stage_cells.append(stage_cell)
                    if payload["coerced_from_text"]:
                        warnings.append({
                            "code": "TEXT_NUMERIC_COERCED",
                            "row": row,
                            "cell": payload["source_cell"],
                            "message": "Khối lượng dạng text được đọc thành Decimal; raw value vẫn được giữ.",
                        })
                    if payload["null_marker"]:
                        warnings.append({
                            "code": "STAGE_NULL_MARKER",
                            "row": row,
                            "cell": payload["source_cell"],
                            "message": "Dấu '-' ở cột giai đoạn được giữ là null, không đổi thành 0.",
                        })
                    if payload["value"] is not None:
                        numeric_stage_cell_count += 1
                        stage_quantities.append(stage_cell)
                        stage_sum += Decimal(payload["value"])

                stage_total = _cell_payload(
                    formula_ws, cached_ws, resolver, row, roles["stage_total"]
                )
                contract_quantity = _cell_payload(
                    formula_ws, cached_ws, resolver, row, roles["contract_quantity"],
                    coerce_numeric_text=True,
                )
                unit_price = _cell_payload(
                    formula_ws, cached_ws, resolver, row, roles["unit_price"],
                    coerce_numeric_text=True,
                )
                amount = _cell_payload(
                    formula_ws, cached_ws, resolver, row, roles["amount"],
                    coerce_numeric_text=True,
                )

                if stage_quantities:
                    stage_allocated_detail_count += 1
                    stage_status = "allocated"
                    stage_bucket_key = None
                    if stage_total["value"] is None:
                        warnings.append({
                            "code": "STAGE_TOTAL_MISSING",
                            "row": row,
                            "cell": stage_total["source_cell"],
                            "message": "Có khối lượng theo giai đoạn nhưng TổNG KHỐI LƯỢNG để trống.",
                        })
                        total_matches = None
                    else:
                        total_matches = Decimal(stage_total["value"]) == stage_sum
                        if not total_matches:
                            warnings.append({
                                "code": "STAGE_TOTAL_MISMATCH",
                                "row": row,
                                "cell": stage_total["source_cell"],
                                "message": "Tổng cột giai đoạn không khớp TỔNG KHỐI LƯỢNG nguồn.",
                            })
                else:
                    unallocated_detail_count += 1
                    stage_status = "unallocated"
                    stage_bucket_key = "unallocated"
                    total_matches = None
                    warnings.append({
                        "code": "DETAIL_WITHOUT_STAGE_QUANTITY",
                        "row": row,
                        "cell": None,
                        "message": "Hạng mục được giữ trong bucket '%s'; không tự phân bổ khối lượng hợp đồng." % UNALLOCATED_STAGE_NAME,
                    })

                def raw_role(name: str) -> Any:
                    column = roles.get(name)
                    return _raw_json_value(formula_ws.cell(row, column).value) if column else None

                lines.append({
                    "kind": "detail",
                    "source_row": row,
                    "source_order": row - header_row,
                    "stt_raw": stt_raw,
                    "stt_cell_type": formula_ws.cell(row, roles["stt"]).data_type,
                    "text_raw": text_raw,
                    "technical_raw": raw_role("technical"),
                    "unit_raw": raw_role("unit"),
                    "category_raw": raw_role("category"),
                    "note_raw": raw_role("note"),
                    "hierarchy_path": _public_hierarchy(parents),
                    "stage_status": stage_status,
                    "stage_bucket_key": stage_bucket_key,
                    "stage_bucket_name": UNALLOCATED_STAGE_NAME if stage_bucket_key else None,
                    "stage_cells": stage_cells,
                    "stage_quantities": stage_quantities,
                    "stage_quantity_sum": _canonical_decimal(stage_sum) if stage_quantities else None,
                    "stage_total": stage_total,
                    "stage_total_matches_sum": total_matches,
                    "contract_quantity": contract_quantity,
                    "unit_price": unit_price,
                    "amount": amount,
                })
                continue

            if _is_total_label(stt_value, description_value):
                totals.append({
                    "kind": "total",
                    "source_row": row,
                    "source_order": row - header_row,
                    "stt_raw": stt_raw,
                    "text_raw": text_raw,
                    "amount": _cell_payload(
                        formula_ws, cached_ws, resolver, row, roles["amount"],
                        coerce_numeric_text=True,
                    ),
                    "note_raw": _raw_json_value(
                        formula_ws.cell(row, roles["note"]).value
                    ) if "note" in roles else None,
                })
                continue

            if text_raw is not None and str(text_raw).strip():
                heading_rows.append(row)
                level = _heading_level(stt_raw, hierarchy_stack)
                while hierarchy_stack and int(hierarchy_stack[-1]["level"]) >= level:
                    hierarchy_stack.pop()
                heading_entry = {
                    "source_row": row,
                    "level": level,
                    "stt_raw": stt_raw,
                    "text_raw": text_raw,
                    "token_kind": _token_kind(stt_raw),
                }
                hierarchy_stack.append(heading_entry)
                lines.append({
                    "kind": "heading",
                    "source_row": row,
                    "source_order": row - header_row,
                    "stt_raw": stt_raw,
                    "stt_cell_type": formula_ws.cell(row, roles["stt"]).data_type,
                    "text_raw": text_raw,
                    "heading_level": level,
                    "hierarchy_path": _public_hierarchy(hierarchy_stack),
                    # Preserve anomalous source cells such as a heading with L=0
                    # or N=0 without mistaking the row for a detail line.
                    "stage_total": _cell_payload(
                        formula_ws, cached_ws, resolver, row, roles["stage_total"]
                    ),
                    "contract_quantity": _cell_payload(
                        formula_ws, cached_ws, resolver, row, roles["contract_quantity"],
                        coerce_numeric_text=True,
                    ),
                })
                continue

            ignored_rows.append(row)
            warnings.append({
                "code": "UNCLASSIFIED_ROW",
                "row": row,
                "cell": None,
                "message": "Dòng có dữ liệu nhưng không có tên hạng mục; không được tự import.",
            })

        details = [line for line in lines if line["kind"] == "detail"]
        duplicate_groups = []
        by_exact_text: Dict[Any, List[int]] = defaultdict(list)
        for line in details:
            by_exact_text[line["text_raw"]].append(line["source_row"])
        for raw_text, rows in by_exact_text.items():
            if len(rows) > 1:
                duplicate_groups.append({"text_raw": raw_text, "source_rows": rows})
                warnings.append({
                    "code": "DUPLICATE_ITEM_TEXT",
                    "row": rows[0],
                    "rows": rows,
                    "cell": None,
                    "message": "Tên hạng mục trùng được giữ nguyên theo source_row, không dedupe.",
                })

        for coordinate, reason in sorted(resolver.unsupported.items()):
            warnings.append({
                "code": "FORMULA_UNSUPPORTED_USING_CACHE",
                "row": formula_ws[coordinate].row,
                "cell": coordinate,
                "message": "Công thức ngoài whitelist; dùng cached value nếu có (%s)." % reason,
            })

        line_rows = [line["source_row"] for line in lines]
        total_rows = [line["source_row"] for line in totals]
        counts = {
            "sheet_row_count": formula_ws.max_row,
            "meaningful_data_row_count": meaningful_rows,
            "line_count": len(lines),
            "heading_count": len(heading_rows),
            "detail_count": len(detail_rows),
            "total_count": len(totals),
            "ignored_row_count": len(ignored_rows),
            "stage_count": len(stage_indexes),
            "stage_definition_count_including_bucket": len(stages),
            "nonblank_stage_cell_count": nonblank_stage_cell_count,
            "stage_allocation_count": numeric_stage_cell_count,
            "stage_allocated_detail_count": stage_allocated_detail_count,
            "unallocated_detail_count": unallocated_detail_count,
            "formula_count": len(formula_cells),
            "formula_cache_missing_count": len(missing_formula_cache),
            "duplicate_item_group_count": len(duplicate_groups),
            "warning_count": len(warnings),
        }
        invariants = {
            "header_detected": True,
            "stage_columns_contiguous": stage_indexes == list(
                range(stage_indexes[0], stage_indexes[-1] + 1)
            ),
            "detail_signature_complete": len(detail_rows) == len(details),
            "totals_excluded_from_lines": not bool(set(line_rows) & set(total_rows)),
            "row_order_preserved": line_rows == sorted(line_rows) and total_rows == sorted(total_rows),
            "formula_cache_complete": not missing_formula_cache,
            "stage_total_and_contract_quantity_are_separate": (
                roles["stage_total"] != roles["contract_quantity"]
            ),
            "partition_counts_match": (
                len(detail_rows) + len(heading_rows) + len(totals) + len(ignored_rows)
                == meaningful_rows
            ),
            "duplicate_items_preserved": sum(len(group["source_rows"]) for group in duplicate_groups)
            == sum(1 for line in details if len(by_exact_text[line["text_raw"]]) > 1),
        }

        def column_descriptor(column: int) -> Dict[str, Any]:
            return {
                "source_column": get_column_letter(column),
                "source_index": column,
                "raw_header": _raw_json_value(formula_ws.cell(header_row, column).value),
            }

        return {
            "ok": True,
            "parser_version": PARSER_VERSION,
            "source": {
                "kind": source_kind,
                "name": source_name,
                "byte_size": len(data),
                "sha256": digest,
                "sheet_names": list(formula_wb.sheetnames),
                "sheet_name": selected_sheet,
                "header_row": header_row,
            },
            "headers": headers,
            "columns": {
                **{role: column_descriptor(column) for role, column in roles.items()},
                "stages": [column_descriptor(column) for column in stage_indexes],
            },
            "stages": stages,
            "counts": counts,
            "invariants": invariants,
            "warnings": warnings,
            "duplicate_item_groups": duplicate_groups,
            "lines": lines,
            "totals": totals,
            "ignored_rows": ignored_rows,
        }
    finally:
        formula_wb.close()
        cached_wb.close()


# Short public alias for future wiring; both names remain pure preview calls.
preview_project_profile = preview_project_profile_xlsx


__all__ = [
    "ProjectProfileImportError",
    "UNALLOCATED_STAGE_NAME",
    "preview_project_profile",
    "preview_project_profile_xlsx",
]
