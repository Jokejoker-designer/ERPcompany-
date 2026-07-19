# -*- coding: utf-8 -*-
"""WO-23 B2/B3/B4 — HOA DON DAU VAO (mua vao): parser + phan loai tat dinh + import 2 pha.

Ghi vao CHINH `hoa_don`(chieu='mua_vao') + `hoa_don_dong` (KHONG tao bang purchase moi).
- Chi thiet_bi / vat_tu -> item_cost_history + stock_ledger(nhap_mua) + cot gia von catalog.
- KHONG dung `gia_gan_nhat` (gia BAN cua WO-15) — mua_vao chi cap nhat cot gia_von_*.
- Tat dinh, KHONG AI. Ca map mo -> match_status='pending' -> popup nguoi duyet.
Nguon that: hd_dau_vao_T1_7\\DS HOA DON MUA VAO ... .xlsx (sheet Smart_KTSC_OK, 17 cot).
"""
import io
import os
import re
import secrets
import time
import unicodedata
from datetime import date, datetime

import db as D

# ---------- chuan hoa ----------
def _norm(s):
    r = "".join(c for c in unicodedata.normalize("NFD", str(s or ""))
                if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", r.replace("đ", "d").replace("Đ", "D").upper()).strip()


def _num(v):
    if v in (None, ""):
        return 0.0
    try:
        return float(str(v).replace(",", "").strip())
    except ValueError:
        return 0.0


def _mst(v):
    """MST NCC: giu ca duoi -00x (chi nhanh) -> chi lay chu so + dau -."""
    return re.sub(r"[^0-9\-]", "", str(v or "")).strip("-")


def _date(v):
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v or "")
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        return m.group(0)
    m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", s)
    if m:
        return "%s-%02d-%02d" % (m.group(3), int(m.group(2)), int(m.group(1)))
    return None


# ---------- B2: alias cot (dò header linh hoat) ----------
COL_ALIAS = {
    "ma_hd": ["SR_HD", "MA_HD", "MAHD", "SERIES"],
    "so_hd": ["SO_HD", "SOHD", "SO HOA DON"],
    "ngay": ["NGAY_HD", "NGAYHD", "NGAY HOA DON", "NGAY"],
    "ncc": ["KHACH HANG", "NHA CUNG CAP", "NCC", "TEN NCC", "TEN DON VI", "TENDONVI"],
    "dia_chi": ["DIA CHI", "DIACHI"],
    "mst": ["MA SO THUE", "MST", "MASOTHUE"],
    "ten_hang": ["TENDM", "MATHANG", "TEN HANG HOA", "TENHANGHOA", "TEN HANG"],
    "dvt": ["DONVI", "DVT", "DON VI"],
    "so_luong": ["LUONG", "SO LUONG", "SOLUONG"],
    "don_gia": ["DGVND", "DON GIA", "DONGIA"],
    "thanh_tien": ["TTVND", "THANH TIEN", "THANHTIEN"],
    "thue_suat": ["TS_GTGT", "THUE SUAT", "THUESUAT", "TS GTGT"],
    "tien_thue": ["THUEVND", "TIEN THUE", "TIENTHUE"],
}


def _detect_columns(header):
    hnorm = [_norm(h) for h in header]
    idx = {}
    for canon, aliases in COL_ALIAS.items():
        for a in aliases:
            an = _norm(a)
            for i, h in enumerate(hnorm):
                if h == an:
                    idx[canon] = i
                    break
            if canon in idx:
                break
    return idx


# ---------- B3: phan loai cost_type + stock_impact (tat dinh, thu tu tren->duoi) ----------
_HANG = ["DAIKIN", "MIDEA", "MITSUBISHI", "TOSHIBA", "PANASONIC", "LG", "GREE",
         "REETECH", "NAGAKAWA", "SAMSUNG", "HITACHI", "FUNIKI", "CASPER", "SUMIKURA"]
# (cost_type, [keyword tren ten da _norm], stock_impact) — khop dau tien thang
COST_RULES = [
    ("thiet_bi", ["MAY LANH", "DIEU HOA", "MAY NEN", "DAN NONG", "DAN LANH", "AHU",
                  "CHILLER", "FCU", "VRV", "VRF", "MULTI", "AM TRAN", "TREO TUONG",
                  "TU DUNG", "AP TRAN", "CASSETTE", "MAY LOC"], 1),
    ("vat_tu", ["ONG DONG", "DAY DIEN", "DAY DAN", "GAS R", "MOI CHAT", "BAO ON",
                "GEN ", "APTOMAT", "APTOMAT", "MCCB", "MCB", " CB ", "ONG PVC",
                "ONG THOAT", "ONG NUOC", "MANG NHUA", "CAO SU", "THECHER", "TAC KE",
                "VIS", "OC VIT", "GIA DO", "TY REN", "MANG", "COV", "DUONG ONG"], 1),
    ("nhan_cong_thue_ngoai", ["NHAN CONG", "CONG LAP", "CONG THAO", "CONG VE SINH",
                              "CONG BAO DUONG", "CONG BAO TRI", "THUE NGOAI", "GIA CONG"], 0),
    ("van_chuyen", ["VAN CHUYEN", "CUOC VAN", "BOC XEP", "SHIP", "CUOC"], 0),
    ("dich_vu", ["PHI DICH VU", "PHI NGAN HANG", "HOC PHI", "BAO DUONG XE", "PHI "], 0),
]


def phan_loai_cost_type(ten):
    n = _norm(ten)
    # thiet bi neu co hang + model-ish
    for ct, kws, stock in COST_RULES:
        for kw in kws:
            if kw.strip() and kw in n:
                return ct, stock
    if any(h in n for h in _HANG) and re.search(r"\d+\s*HP|\d+\s*KW|\d+\s*BTU|MODEL", n):
        return "thiet_bi", 1
    return "khac", 0


def _extract_brand(n):
    for h in _HANG:
        if h in n:
            return h
    return ""


def _extract_power(n):
    m = re.search(r"(\d+(?:[.,]\d+)?)\s*(HP|KW|BTU)", n)
    return (m.group(1).replace(",", ".") + m.group(2)) if m else ""


def _extract_model(n):
    m = re.search(r"\b([A-Z]{2,}[A-Z0-9\-]*\d[A-Z0-9\-]*)\b", n)
    return m.group(1) if m else ""


def _item_key(ten, dvt=""):
    """Khoa chuan hoa: NHOM|HANG|CONGSUAT|MODEL. KHONG dua DVT vao khoa
    (DVT hay thieu/khac nhau -> lam lech khop). DVT chi tinh diem tin cay."""
    n = _norm(ten)
    ct, _ = phan_loai_cost_type(ten)
    parts = [ct, _extract_brand(n), _extract_power(n), _extract_model(n)]
    key = "|".join(p for p in parts if p)
    return key or n[:60]


# ---------- B4: diem tin cay khop item (contract §4) ----------
def _match(conn, ten, dvt, cost_type):
    """Tra (item_key, confidence, status). Chi thiet_bi/vat_tu can khop item;
    con lai = chi phi -> auto (khong can dinh danh ton kho)."""
    if cost_type not in ("thiet_bi", "vat_tu"):
        return _item_key(ten, dvt), 1.0, "auto"
    n = _norm(ten)
    # 1) alias rule (tu hoc) — chinh xac +0.50
    rule = conn.execute("""SELECT item_key FROM item_alias_rule
        WHERE is_active=1 AND ?=UPPER(alias_text) ORDER BY priority LIMIT 1""", (n,)).fetchone()
    base = 0.50 if rule else 0.0
    ikey = rule["item_key"] if rule else _item_key(ten, dvt)
    score = base
    if _extract_brand(n):
        score += 0.15
    if _extract_model(n):
        score += 0.25
    if _extract_power(n):
        score += 0.15
    if _norm(dvt):
        score += 0.10
    score += 0.10  # co nhom (thiet_bi/vat_tu)
    score = min(score, 1.0)
    status = "auto" if score >= 0.85 else ("pending" if score >= 0.60 else "unmatched")
    return ikey, round(score, 2), status


# ---------- Parser file ----------
def parse_file(path_or_bytes, filename=""):
    import openpyxl
    src = io.BytesIO(path_or_bytes) if isinstance(path_or_bytes, (bytes, bytearray)) else path_or_bytes
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return {"hoa_don": {}, "loi": "File rong"}
    header = rows[0]
    idx = _detect_columns(header)
    missing = [k for k in ("ncc", "ten_hang", "don_gia") if k not in idx]
    if missing:
        return {"hoa_don": {}, "loi": "Khong nhan cot: " + ", ".join(missing)}

    def g(row, col):
        i = idx.get(col)
        return row[i] if i is not None and i < len(row) else None

    groups = {}  # (ma_hd, ngay, mst) -> {header, dong[]}
    for row in rows[1:]:
        if not any(row):
            continue
        ma_hd = str(g(row, "ma_hd") or "").strip()
        ngay = _date(g(row, "ngay"))
        mst = _mst(g(row, "mst"))
        ten = str(g(row, "ten_hang") or "").strip()
        sl = _num(g(row, "so_luong"))
        dg = _num(g(row, "don_gia"))
        tt = _num(g(row, "thanh_tien"))
        # dong mo ta rong (LUONG=0 & DGVND=0 & TTVND=0) -> bo qua
        if sl == 0 and dg == 0 and tt == 0:
            continue
        if not ten and not tt:
            continue
        key = (ma_hd, ngay, mst)
        grp = groups.setdefault(key, {
            "ma_hd": ma_hd, "so_hd": str(g(row, "so_hd") or "").strip(), "ngay": ngay,
            "ncc": str(g(row, "ncc") or "").strip(), "mst": mst,
            "dia_chi": str(g(row, "dia_chi") or "").strip(), "dong": []})
        ts_raw = _norm(g(row, "thue_suat"))
        vat = 0.0 if ("KCT" in ts_raw or ts_raw == "") else _num(g(row, "thue_suat"))
        grp["dong"].append({
            "ten_hang_hoa": ten, "dvt": str(g(row, "dvt") or "").strip(),
            "so_luong": sl, "don_gia": dg, "thanh_tien": tt or (sl * dg),
            "thue_suat": vat, "tien_thue": _num(g(row, "tien_thue"))})
    return {"hoa_don": groups}


def classify_groups(conn, groups):
    """Gan cost_type/stock_impact/item_key/confidence/status cho tung dong."""
    stat = {"hoa_don": 0, "dong": 0, "auto": 0, "pending": 0, "unmatched": 0,
            "thiet_bi": 0, "vat_tu": 0, "chi_phi": 0}
    for grp in groups.values():
        stat["hoa_don"] += 1
        for d in grp["dong"]:
            ct, stock = phan_loai_cost_type(d["ten_hang_hoa"])
            ikey, conf, status = _match(conn, d["ten_hang_hoa"], d["dvt"], ct)
            d.update({"cost_type": ct, "stock_impact": stock, "item_key": ikey,
                      "match_confidence": conf, "match_status": status})
            stat["dong"] += 1
            stat[status] = stat.get(status, 0) + 1
            if ct == "thiet_bi":
                stat["thiet_bi"] += 1
            elif ct == "vat_tu":
                stat["vat_tu"] += 1
            else:
                stat["chi_phi"] += 1
    return stat


# ---------- 2 pha: preview / commit ----------
_TOKENS = {}  # token -> {"groups":..., "filename":..., "het_han": epoch}


def import_preview(conn, file_bytes, filename):
    parsed = parse_file(file_bytes, filename)
    if parsed.get("loi"):
        return {"ok": False, "error": parsed["loi"]}
    groups = parsed["hoa_don"]
    stat = classify_groups(conn, groups)
    # danh dau trung (da co trong hoa_don mua_vao)
    trung = 0
    for grp in groups.values():
        dup = conn.execute("""SELECT id FROM hoa_don WHERE ma_hd=? AND ngay=? AND mst=?
            AND chieu='mua_vao'""", (grp["ma_hd"], grp["ngay"], grp["mst"])).fetchone()
        grp["_trung"] = bool(dup)
        if dup:
            trung += 1
    tok = "hdmv_" + secrets.token_urlsafe(12)
    _TOKENS[tok] = {"groups": groups, "filename": filename, "het_han": time.time() + 600}
    hd_list = [{"ma_hd": g["ma_hd"], "so_hd": g["so_hd"], "ngay": g["ngay"], "ncc": g["ncc"],
                "mst": g["mst"], "so_dong": len(g["dong"]),
                "tong": sum(d["thanh_tien"] + d["tien_thue"] for d in g["dong"]),
                "trung": g["_trung"]} for g in groups.values()]
    dong_list = [dict(d, so_hd=g["so_hd"]) for g in groups.values() for d in g["dong"]]
    return {"ok": True, "confirm_token": tok, "filename": filename,
            "summary": stat, "so_trung": trung,
            "hoa_don": hd_list, "dong": dong_list[:500],
            "so_pending": stat.get("pending", 0) + stat.get("unmatched", 0)}


def import_commit(conn, sess, token, overrides=None):
    from api_write import audit
    entry = _TOKENS.pop(token, None)
    if not entry:
        raise ValueError("Token het han hoac da dung — lam lai buoc xem truoc.")
    if entry["het_han"] < time.time():
        raise ValueError("Token het han (10 phut).")
    groups = entry["groups"]
    fname = entry["filename"]
    # ap override: {so_hd,ten,cost_type?,item_key?,stock_impact?,match_status?}
    ov_map = {}
    for o in (overrides or []):
        ov_map[(o.get("so_hd"), o.get("ten"))] = o

    st = {"hoa_don_moi": 0, "trung_bo_qua": 0, "dong": 0, "cost_rows": 0, "stock_rows": 0,
          "pending_con": 0}
    for grp in groups.values():
        dup = conn.execute("""SELECT id FROM hoa_don WHERE ma_hd=? AND ngay=? AND mst=?
            AND chieu='mua_vao'""", (grp["ma_hd"], grp["ngay"], grp["mst"])).fetchone()
        if dup:
            st["trung_bo_qua"] += 1
            continue
        tong_tt = sum(d["thanh_tien"] for d in grp["dong"])
        tong_thue = sum(d["tien_thue"] for d in grp["dong"])
        conn.execute("""INSERT INTO hoa_don(ma_hd, ngay, customer_id, mst, ten_don_vi, dia_chi,
            tong_truoc_thue, tong_thue, tong_cong, chieu, nguon_file)
            VALUES(?,?,?,?,?,?,?,?,?,?,?)""",
            (grp["ma_hd"], grp["ngay"], None, grp["mst"], grp["ncc"], grp["dia_chi"],
             tong_tt, tong_thue, tong_tt + tong_thue, "mua_vao", fname))
        hd_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        st["hoa_don_moi"] += 1
        for d in grp["dong"]:
            o = ov_map.get((grp["so_hd"], d["ten_hang_hoa"]))
            cost_type = (o.get("cost_type") if o else None) or d["cost_type"]
            item_key = (o.get("item_key") if o else None) or d["item_key"]
            stock_impact = o.get("stock_impact") if (o and o.get("stock_impact") is not None) \
                else d["stock_impact"]
            status = (o.get("match_status") if o else None) or d["match_status"]
            if o:
                status = "confirmed"  # nguoi duyet da sua -> chot
            conn.execute("""INSERT INTO hoa_don_dong(hoa_don_id, ten_hang_hoa, dvt, so_luong,
                don_gia, thanh_tien, thue_suat, tien_thue, cost_type, stock_impact, item_key,
                match_confidence, match_status)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (hd_id, d["ten_hang_hoa"], d["dvt"], d["so_luong"], d["don_gia"], d["thanh_tien"],
                 d["thue_suat"], d["tien_thue"], cost_type, stock_impact, item_key,
                 d["match_confidence"], status))
            dong_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
            st["dong"] += 1
            if status == "pending" or status == "unmatched":
                st["pending_con"] += 1
            # chi thiet_bi/vat_tu -> gia von + ton (bo qua neu con pending chua chot)
            if stock_impact and status in ("auto", "confirmed"):
                cost_with_vat = d["thanh_tien"] + d["tien_thue"]
                conn.execute("""INSERT INTO item_cost_history(item_key, item_name, item_group,
                    brand, model, uom, supplier_name, supplier_mst, hoa_don_id, hoa_don_dong_id,
                    purchase_date, quantity, unit_cost, vat_rate, cost_with_vat, source_type)
                    VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (item_key, d["ten_hang_hoa"], cost_type, _extract_brand(_norm(d["ten_hang_hoa"])),
                     _extract_model(_norm(d["ten_hang_hoa"])), d["dvt"], grp["ncc"], grp["mst"],
                     hd_id, dong_id, grp["ngay"], d["so_luong"], d["don_gia"], d["thue_suat"],
                     cost_with_vat, "mua_vao"))
                st["cost_rows"] += 1
                conn.execute("""INSERT INTO stock_ledger(item_key, item_name, movement_type,
                    source_type, source_id, source_line_id, movement_date, qty_in, unit_cost, amount, note)
                    VALUES(?,?,?,?,?,?,?,?,?,?,?)""",
                    (item_key, d["ten_hang_hoa"], "nhap_mua", "mua_vao", hd_id, dong_id,
                     grp["ngay"], d["so_luong"], d["don_gia"], d["thanh_tien"],
                     "Nhap mua tu " + grp["ma_hd"]))
                st["stock_rows"] += 1
                # cap nhat cot GIA VON catalog — KHONG dung gia_gan_nhat (gia ban)
                _cap_nhat_gia_von_catalog(conn, item_key, d, grp)
    audit(conn, sess, "import_mua", "hoa_don", "batch",
          "Import HD dau vao: %d HD, %d dong, %d cost, %d ton, %d pending" %
          (st["hoa_don_moi"], st["dong"], st["cost_rows"], st["stock_rows"], st["pending_con"]))
    conn.commit()
    return st


def _cap_nhat_gia_von_catalog(conn, item_key, d, grp):
    """Cap nhat gia_von_gan_nhat/gia_von_tb tren mat_hang_tu_hoa_don theo item_key.
    KHONG dung gia_gan_nhat (gia ban). Tao dong catalog neu chua co (theo ten)."""
    row = conn.execute("SELECT id, gia_von_tb, so_lan_ban FROM mat_hang_tu_hoa_don WHERE ten_hang_hoa=?",
                       (d["ten_hang_hoa"],)).fetchone()
    if row:
        old_tb = row["gia_von_tb"] or d["don_gia"]
        new_tb = (old_tb + d["don_gia"]) / 2 if row["gia_von_tb"] else d["don_gia"]
        conn.execute("""UPDATE mat_hang_tu_hoa_don SET item_key=?, item_group=?,
            gia_von_gan_nhat=?, gia_von_tb=?, ncc_gan_nhat=?, ngay_mua_gan_nhat=? WHERE id=?""",
            (item_key, d["cost_type"], d["don_gia"], new_tb, grp["ncc"], grp["ngay"], row["id"]))
    else:
        conn.execute("""INSERT INTO mat_hang_tu_hoa_don(ten_hang_hoa, dvt, item_key, item_group,
            gia_von_gan_nhat, gia_von_tb, ncc_gan_nhat, ngay_mua_gan_nhat)
            VALUES(?,?,?,?,?,?,?,?)""",
            (d["ten_hang_hoa"], d["dvt"], item_key, d["cost_type"],
             d["don_gia"], d["don_gia"], grp["ncc"], grp["ngay"]))
