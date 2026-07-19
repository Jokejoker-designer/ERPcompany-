# -*- coding: utf-8 -*-
"""Regression tests for P1 finding FIND-007 (WO: harden /api/import_run's
client-supplied path/paths before they reach openpyxl.load_workbook()).

Covers the fixes applied to import_excel.py (_find_file/import_customers/
import_invoices) and server.py (_import_run):
  1. UNC-style paths (`\\\\host\\share\\...`) are rejected outright, in
     `path` and in any entry of `paths`.
  2. Non-.xlsx paths are rejected before _find_file's filesystem lookup.
  3. Oversized files are rejected before openpyxl.load_workbook() is called
     (IMPORT_MAX_BYTES cap, monkeypatched small for a fast test).
  4. Preview-mode (no `commit`) /api/import_run calls for loai=customers /
     invoices / invoices_mua now write an audit_log row (previously silent).

Runs entirely against throwaway in-memory SQLite DBs built from schema.sql
and throwaway files under a tempfile.TemporaryDirectory(). NEVER opens,
reads, or writes data/thanh_hoai.db, and NEVER touches any real file under
D:\\2025 or D:\\2026. NEVER starts server.py's HTTP server (server module is
only imported -- main()/ThreadingHTTPServer is behind `if __name__ ==
"__main__"` and is never invoked here) and makes no HTTP requests.

Run with:
    python -m unittest test_import_run_regression -v
"""
import os
import sqlite3
import tempfile
import unittest
from unittest import mock

import openpyxl

import api_write as AW
import import_excel as IE
import server as SRV

APP_ROOT = os.path.dirname(os.path.abspath(__file__))
SCHEMA_PATH = os.path.join(APP_ROOT, "schema.sql")

UNC_PATH = r"\\evil-share\public\Customer data.xlsx"
UNC_PATH_SLASH = r"//evil-share/public/Customer data.xlsx"
# WO27 (FIND-007 reopened): mixed-separator UNC forms that GetFullPathNameW
# still normalizes to a genuine \\host\share UNC path on Windows, but which
# the old startswith("\\\\")/startswith("//") check let slip through.
UNC_PATH_MIXED_FS_BS = "/\\evil-share\\public\\Customer data.xlsx"
UNC_PATH_MIXED_BS_FS = "\\/evil-share\\public\\Customer data.xlsx"


def _add_col_if_missing(conn, table, column, declaration):
    if column not in {r["name"] for r in conn.execute("PRAGMA table_info(%s)" % table)}:
        conn.execute("ALTER TABLE %s ADD COLUMN %s %s" % (table, column, declaration))


def _make_conn():
    """Fresh in-memory DB from the real schema.sql (structure only, no data).
    Mirrors test_authz_regression.py's _make_conn()."""
    conn = sqlite3.connect(":memory:")
    conn.row_factory = sqlite3.Row
    with open(SCHEMA_PATH, "r", encoding="utf-8") as f:
        conn.executescript(f.read())
    _add_col_if_missing(conn, "cong_viec_ktv", "quotation_id", "INTEGER")
    _add_col_if_missing(conn, "cong_viec_ktv", "ktv_id", "INTEGER")
    _add_col_if_missing(conn, "quotation", "trang_thai_doi_chieu", "TEXT")
    _add_col_if_missing(conn, "customer", "nguon", "TEXT")
    _add_col_if_missing(conn, "customer", "ghi_chu", "TEXT")
    conn.commit()
    return conn


class _NoCloseConn:
    """Proxy a real sqlite3 connection so a caller's conn.close() is a
    no-op -- lets a test inspect DB state (e.g. audit_log) after the
    handler under test believes it has closed the connection."""

    def __init__(self, real):
        self._real = real

    def close(self):
        pass  # swallow -- keep self._real open for post-call assertions

    def __getattr__(self, name):
        return getattr(self._real, name)


class _FakeHandler:
    """Minimal stand-in for server.Handler -- only implements the one
    method (_send_json) that Handler._import_run touches on `self`.
    Calling Handler._import_run(fake, sess, body) as an unbound method
    never constructs a real BaseHTTPRequestHandler, never binds a socket,
    and never starts the app's HTTP server."""

    def __init__(self):
        self.status = None
        self.payload = None

    def _send_json(self, obj, status=200, set_cookie=None):
        self.status = status
        self.payload = obj


def _write_xlsx(path, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    wb.save(path)


CUSTOMER_HEADER = ["Ma", "TenKhachHang", "NguoiLienHe", "MST", "Email", "SDT", "GhiChu"]


class FindSevenFindFileHardeningTest(unittest.TestCase):
    """FIND-007 items 1-3: _find_file must reject UNC paths, non-.xlsx
    paths and oversized files -- all BEFORE openpyxl.load_workbook() is
    ever reached -- for both import_customers (single `path`) and
    import_invoices (`paths` list, incl. a bad entry mixed with a good
    one)."""

    def setUp(self):
        self.conn = _make_conn()
        self.tmpdir = tempfile.TemporaryDirectory()

    def tearDown(self):
        self.conn.close()
        self.tmpdir.cleanup()

    # ---- 1. UNC path rejection -------------------------------------
    def test_find_file_rejects_unc_path(self):
        with mock.patch("openpyxl.load_workbook") as mock_lw:
            with self.assertRaises(AW.ValidationError):
                IE._find_file(UNC_PATH)
            mock_lw.assert_not_called()

    def test_import_customers_rejects_unc_path(self):
        with mock.patch("openpyxl.load_workbook") as mock_lw:
            with self.assertRaises(AW.ValidationError):
                IE.import_customers(UNC_PATH, commit=False, conn=self.conn)
            mock_lw.assert_not_called()

    def test_import_invoices_rejects_unc_entry_in_paths_list(self):
        # A UNC entry anywhere in `paths` must be rejected, not just a
        # lone `path` string. First entry is a real, legit, unmocked
        # .xlsx (proves the good entry is unaffected / still processed
        # normally); the second (UNC) entry must still be rejected.
        good = os.path.join(self.tmpdir.name, "Invoice_ok.xlsx")
        _write_xlsx(good, [["MaHD", "NgayHoaDon", "MaSoThue", "TenDonVi"]])
        with self.assertRaises(AW.ValidationError):
            IE.import_invoices([good, UNC_PATH], commit=False, conn=self.conn)
        # Isolate the bad entry alone to prove it specifically never
        # reaches openpyxl.load_workbook.
        with mock.patch("openpyxl.load_workbook") as mock_lw:
            with self.assertRaises(AW.ValidationError):
                IE.import_invoices([UNC_PATH], commit=False, conn=self.conn)
            mock_lw.assert_not_called()

    # ---- 1b. WO27: mixed-separator UNC bypass (FIND-007 reopened) ---
    # The old check only matched startswith("\\\\") / startswith("//") --
    # i.e. 2 leading separators of the SAME kind. A path starting with
    # one kind then the other ("/\\..." or "\\/...") slipped straight
    # through to the isfile/fuzzy-lookup branch. Windows' own
    # GetFullPathNameW normalizes both of these into a real \\host\share
    # UNC path, confirming the bypass is a genuine SMB/SSRF risk, not
    # just a cosmetic string-matching gap. Fixed via
    # re.match(r'^[/\\]{2,}', path).
    def test_find_file_rejects_unc_path_pure_forward_slash(self):
        # Non-regression companion to test_find_file_rejects_unc_path:
        # the pure "//" form must keep being rejected too.
        with mock.patch("openpyxl.load_workbook") as mock_lw:
            with self.assertRaises(AW.ValidationError):
                IE._find_file(UNC_PATH_SLASH)
            mock_lw.assert_not_called()

    def test_find_file_rejects_mixed_separator_unc_forward_then_back(self):
        with mock.patch("openpyxl.load_workbook") as mock_lw:
            with self.assertRaises(AW.ValidationError):
                IE._find_file(UNC_PATH_MIXED_FS_BS)
            mock_lw.assert_not_called()

    def test_find_file_rejects_mixed_separator_unc_back_then_forward(self):
        with mock.patch("openpyxl.load_workbook") as mock_lw:
            with self.assertRaises(AW.ValidationError):
                IE._find_file(UNC_PATH_MIXED_BS_FS)
            mock_lw.assert_not_called()

    def test_import_customers_rejects_mixed_separator_unc_path_forward_then_back(self):
        with mock.patch("openpyxl.load_workbook") as mock_lw:
            with self.assertRaises(AW.ValidationError):
                IE.import_customers(UNC_PATH_MIXED_FS_BS, commit=False, conn=self.conn)
            mock_lw.assert_not_called()

    def test_import_invoices_rejects_mixed_separator_unc_entry_back_then_forward(self):
        with mock.patch("openpyxl.load_workbook") as mock_lw:
            with self.assertRaises(AW.ValidationError):
                IE.import_invoices([UNC_PATH_MIXED_BS_FS], commit=False, conn=self.conn)
            mock_lw.assert_not_called()

    # ---- 2. Non-.xlsx extension rejection ---------------------------
    def test_find_file_rejects_non_xlsx_extension(self):
        bad = os.path.join(self.tmpdir.name, "Customer data.csv")
        with open(bad, "w", encoding="utf-8") as f:
            f.write("Ma,TenKhachHang\n1,Test\n")  # real file that DOES exist on disk
        with mock.patch("openpyxl.load_workbook") as mock_lw:
            with self.assertRaises(AW.ValidationError):
                IE._find_file(bad)
            mock_lw.assert_not_called()

    def test_import_customers_rejects_non_xlsx_extension(self):
        bad = os.path.join(self.tmpdir.name, "Customer data.txt")
        with open(bad, "w", encoding="utf-8") as f:
            f.write("not an xlsx file")
        with mock.patch("openpyxl.load_workbook") as mock_lw:
            with self.assertRaises(AW.ValidationError):
                IE.import_customers(bad, commit=False, conn=self.conn)
            mock_lw.assert_not_called()

    # ---- 3. Oversized file rejection ---------------------------------
    def test_find_file_rejects_oversized_file(self):
        big = os.path.join(self.tmpdir.name, "Customer data.xlsx")
        with open(big, "wb") as f:
            f.write(b"0" * 500)  # 500 bytes -- larger than the patched cap below
        with mock.patch.object(IE, "IMPORT_MAX_BYTES", 100), \
             mock.patch("openpyxl.load_workbook") as mock_lw:
            with self.assertRaises(AW.ValidationError):
                IE._find_file(big)
            mock_lw.assert_not_called()

    def test_import_customers_rejects_oversized_file(self):
        big = os.path.join(self.tmpdir.name, "Customer data.xlsx")
        with open(big, "wb") as f:
            f.write(b"0" * 500)
        with mock.patch.object(IE, "IMPORT_MAX_BYTES", 100), \
             mock.patch("openpyxl.load_workbook") as mock_lw:
            with self.assertRaises(AW.ValidationError):
                IE.import_customers(big, commit=False, conn=self.conn)
            mock_lw.assert_not_called()

    # ---- sanity: a legitimate small .xlsx still works (no regression) --
    def test_import_customers_still_accepts_legit_small_xlsx(self):
        good = os.path.join(self.tmpdir.name, "Customer data.xlsx")
        _write_xlsx(good, [CUSTOMER_HEADER, ["1", "Khach Test A", "", "", "", "", ""]])
        out = IE.import_customers(good, commit=False, conn=self.conn)
        self.assertEqual(out["stats"]["them_moi"], 1)


class FindSevenPreviewAuditTest(unittest.TestCase):
    """FIND-007 item 4: /api/import_run must write an audit_log row on
    EVERY call for loai=customers/invoices/invoices_mua, including
    preview (no `commit`) calls -- previously the `if commit or loai ==
    "doichieu"` guard skipped logging entirely for preview reads.

    Exercises server.Handler._import_run directly (unbound-method call
    against a minimal fake `self`) -- never constructs a real
    BaseHTTPRequestHandler, never binds a socket, never starts the app's
    HTTP server, and never touches data/thanh_hoai.db (db.get_conn is
    monkeypatched to a throwaway in-memory connection for the duration
    of each test)."""

    def setUp(self):
        self.conn = _make_conn()
        self.tmpdir = tempfile.TemporaryDirectory()
        self.customer_xlsx = os.path.join(self.tmpdir.name, "Customer data.xlsx")
        _write_xlsx(self.customer_xlsx,
                    [CUSTOMER_HEADER, ["1", "Khach Test A", "", "", "", "", ""]])
        self._patcher = mock.patch.object(
            SRV.D, "get_conn", lambda: _NoCloseConn(self.conn))
        self._patcher.start()

    def tearDown(self):
        self._patcher.stop()
        self.tmpdir.cleanup()
        self.conn.close()

    def _audit_rows(self):
        return self.conn.execute("SELECT * FROM audit_log").fetchall()

    def test_preview_customers_call_is_now_audited(self):
        fake = _FakeHandler()
        sess = {"username": "ketoan.test", "role": "Ke toan"}
        body = {"loai": "customers", "path": self.customer_xlsx}  # no "commit" -> preview
        SRV.Handler._import_run(fake, sess, body)
        self.assertEqual(fake.status, 200, fake.payload)
        rows = self._audit_rows()
        self.assertEqual(
            len(rows), 1,
            "preview-mode import_run (loai=customers, no commit) must write "
            "exactly one audit_log row -- previously wrote none")
        self.assertEqual(rows[0]["hanh_dong"], "import")
        self.assertEqual(rows[0]["bang"], "import_run")
        self.assertEqual(rows[0]["ban_ghi_id"], "customers")
        self.assertEqual(rows[0]["user"], "ketoan.test")
        self.assertEqual(rows[0]["role"], "Ke toan")

    def test_commit_customers_call_still_audited_exactly_once(self):
        # Regression guard: broadening the audit condition must not cause
        # double-logging on the existing commit path.
        fake = _FakeHandler()
        sess = {"username": "ketoan.test", "role": "Ke toan"}
        body = {"loai": "customers", "path": self.customer_xlsx, "commit": True}
        SRV.Handler._import_run(fake, sess, body)
        self.assertEqual(fake.status, 200, fake.payload)
        rows = self._audit_rows()
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["ban_ghi_id"], "customers")

    def test_rejected_unc_path_returns_400_and_writes_no_audit_row(self):
        # Validation failures (FIND-007 fixes) must surface as HTTP 400
        # via the same AW.ValidationError -> 400 convention _export()
        # already uses, and must not reach the audit block at all.
        fake = _FakeHandler()
        sess = {"username": "ketoan.test", "role": "Ke toan"}
        body = {"loai": "customers", "path": UNC_PATH}
        SRV.Handler._import_run(fake, sess, body)
        self.assertEqual(fake.status, 400, fake.payload)
        self.assertEqual(self._audit_rows(), [])


if __name__ == "__main__":
    unittest.main()
