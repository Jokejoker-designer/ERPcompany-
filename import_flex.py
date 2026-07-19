# -*- coding: utf-8 -*-
"""WO-23 B9 — Import LINH HOAT (ban do cot tu hoc) + cau noi tao bao gia.

Giai bai: file NCC/khach KHONG chuan mau -> import co dinh khong doc duoc.
Tat dinh, KHONG AI: nguoi map cot 1 lan -> he nho `import_profile` (theo file_signature)
-> tu ap lan sau. 3 pha: preview (luoi tho + goi y profile) / map (dong co cau truc +
pipeline B3) / commit (ghi theo scope). Cau noi: moi_thau_khach -> tao_bao_gia_tu_list.
"""
import hashlib
import io
import json
import re
import secrets
import time
import unicodedata
from datetime import date, datetime

import db as D
import import_hd_dauvao as HM   # tai dung _norm/_num/_item_key/phan_loai_cost_type/_match

_TOKENS = {}   # token -> {"lines":[...], "scope":..., "target":{...}, "het_han": epoch}


def _norm_header(cells):
    return [HM._norm(c) for c in cells]


def _file_signature(header_cells):
    """Hash chuoi header da chuan hoa -> tu nhan lai file cung mau."""
    key = "|".join(_norm_header(header_cells))
    return hashlib.sha256(key.encode("utf-8")).hexdigest()[:24]


def _load_sheet_rows(file_bytes, sheet=None, max_row=None):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheets = wb.sheetnames
    ws = wb[sheet] if (sheet and sheet in sheets) else wb[sheets[0]]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        rows.append(list(row))
        if max_row and i + 1 >= max_row:
            break
    wb.close()
    return sheets, rows


def import_flex_preview(conn, file_bytes, filename):
    """Tra luoi 30 dong tho + danh sach sheet. Neu file_signature khop -> profile_goi_y."""
    sheets, rows = _load_sheet_rows(file_bytes, max_row=30)
    grid = [[("" if c is None else str(c)) for c in r] for r in rows]
    # doan dong header = dong nhieu o chu nhat trong 8 dong dau
    header_guess = 0
    best = -1
    for i, r in enumerate(rows[:8]):
        n = sum(1 for c in r if isinstance(c, str) and c.strip())
        if n > best:
            best, header_guess = n, i
    profile = None
    if rows:
        sig = _file_signature(rows[header_guess])
        p = conn.execute("SELECT * FROM import_profile WHERE file_signature=? AND is_active=1",
                         (sig,)).fetchone()
        if p:
            profile = {"sheet_name": p["sheet_name"], "header_row": p["header_row"],
                       "data_start_row": p["data_start_row"], "col_map": json.loads(p["col_map"] or "{}"),
                       "scope": p["scope"], "ten_profile": p["ten_profile"], "doi_tac": p["doi_tac"]}
    return {"ok": True, "sheets": sheets, "grid": grid[:30],
            "header_row_goi_y": header_guess, "data_start_goi_y": header_guess + 1,
            "profile_goi_y": profile}


# WO-29: loc dong TIEU DE/TONG KET lan vao vung du lieu — file that thuong co dong "CONG"/
# "VAT 8%"/"TONG" o CUOI bang tinh tong (khong phai hang muc/hang hoa thuc), hoac dong dau
# lap lai chinh ten cot header (vd chon nham data_start_row som 1 dong). Ap dung cho MOI
# scope vi tat ca deu doc 1 cot van ban chinh lam "ten_hang"/hang_muc.
_DONG_BO_QUA = {"NOI DUNG", "HANG MUC", "TEN HANG", "MO TA", "CONG", "TONG", "TONG CONG",
               "THANH TIEN", "DON GIA", "SO LUONG", "STT", "KHOI LUONG", "DVT", "GHI CHU"}


def _la_dong_bo_qua(ten):
    t = HM._norm(ten).strip()
    if not t:
        return True
    if t in _DONG_BO_QUA:
        return True
    if t.startswith("TONG") or t.startswith("THANH TIEN"):
        return True
    if re.fullmatch(r"VAT\s*(?:8|10)?\s*%?", t) or t.startswith("THUE VAT"):   # khong nham "VAT TU" voi VAT
        return True
    return False


def import_flex_map(conn, params):
    """Doc theo {sheet, header_row, data_start_row, col_map} -> dong co cau truc.
    WO-29 Phase 1: tach nhanh theo scope — cac scope "hang hoa" (moi_thau_khach/
    bao_gia_ncc/hoa_don_dau_vao) GIU NGUYEN pipeline B3 (khop item/gia von); scope
    bbnt_cu/pxk_cu la CHUNG TU KHAC HINH DANG (khong co cot_type/item_key trong schema
    dich) -> vai tro cot RIENG, khong qua pipeline khop hang hoa."""
    file_bytes = params["_file_bytes"]
    sheet = params.get("sheet")
    header_row = int(params.get("header_row", 0))
    data_start = int(params.get("data_start_row", header_row + 1))
    col_map = params.get("col_map") or {}
    scope = params.get("scope") or "moi_thau_khach"
    target = params.get("target") or {}
    _, rows = _load_sheet_rows(file_bytes, sheet=sheet)
    if header_row >= len(rows):
        return {"ok": False, "error": "Dòng header vượt quá số dòng."}

    def cell(row, key):
        idx = col_map.get(key)
        if idx is None or idx == "":
            return None
        idx = int(idx)
        return row[idx] if idx < len(row) else None

    def co_du_lieu(row):
        return any(v is not None and str(v).strip() for v in row)

    lines = []
    last_vat = None
    if scope == "bbnt_cu":
        # bbnt_item: hang_muc/don_gia/thanh_tien/kl_hop_dong/kl_thuc_te/ket_qua/ghi_chu —
        # KHONG co so_luong/dvt/item_key (khong phai chung tu hang hoa co catalog).
        for row in rows[data_start:]:
            if not co_du_lieu(row):
                continue
            hang_muc = str(cell(row, "ten_hang") or "").strip()   # dung chung role "ten_hang"
            if not hang_muc or _la_dong_bo_qua(hang_muc):          # (= "Hang muc" o nhan BBNT)
                continue
            lines.append({"ten_hang": hang_muc, "hang_muc": hang_muc,
                          "don_gia": HM._num(cell(row, "don_gia")) or 0,
                          "thanh_tien": HM._num(cell(row, "thanh_tien")) or 0,
                          "kl_hop_dong": str(cell(row, "kl_hop_dong") or "").strip(),
                          "kl_thuc_te": str(cell(row, "kl_thuc_te") or "").strip(),
                          "ket_qua": str(cell(row, "ket_qua") or "").strip(),
                          "ghi_chu": str(cell(row, "ghi_chu") or "").strip(),
                          "match_status": "auto", "confidence": 1.0})
    elif scope == "pxk_cu":
        # pxk_dong: ten_hang/dvt/so_luong/ghi_chu — KHONG co don_gia (khong phai chung tu tien).
        for row in rows[data_start:]:
            if not co_du_lieu(row):
                continue
            ten = str(cell(row, "ten_hang") or "").strip()
            if not ten or _la_dong_bo_qua(ten):
                continue
            model = str(cell(row, "model") or "").strip()
            lines.append({"ten_hang": ten, "model": model,
                          "dvt": str(cell(row, "dvt") or "").strip(),
                          "so_luong": HM._num(cell(row, "so_luong")) or 1,
                          "ghi_chu": str(cell(row, "ghi_chu") or "").strip(),
                          "match_status": "auto", "confidence": 1.0})
    else:  # moi_thau_khach / bao_gia_ncc / hoa_don_dau_vao — pipeline B3 CU, KHONG DOI
        for row in rows[data_start:]:
            if not co_du_lieu(row):
                continue
            ten = str(cell(row, "ten_hang") or "").strip()
            if not ten or _la_dong_bo_qua(ten):
                continue
            model = str(cell(row, "model") or "").strip()
            dvt = str(cell(row, "dvt") or "").strip()
            sl = HM._num(cell(row, "so_luong")) or 1
            dg = HM._num(cell(row, "don_gia"))
            thanh_tien = HM._num(cell(row, "thanh_tien"))
            ts_raw = HM._norm(cell(row, "thue_suat"))
            vat_match = re.search(r"(\d+(?:\.\d+)?)\s*%", ts_raw)
            if vat_match:
                vat = float(vat_match.group(1))
                last_vat = vat
            elif "KCT" in ts_raw:
                vat = 0.0
            elif ts_raw == "" and last_vat is not None:
                vat = last_vat
            else:
                vat = HM._num(cell(row, "thue_suat")) or 0.0
            full_name = (ten + " " + model).strip()
            ct, stock = HM.phan_loai_cost_type(full_name)
            ikey, conf, status = HM._match(conn, full_name, dvt, ct)
            lines.append({"ten_hang": ten, "model": model, "dvt": dvt, "so_luong": sl,
                          "don_gia": dg, "thanh_tien": thanh_tien, "thue_suat": vat,
                          "item_key": ikey, "cost_type": ct,
                          "stock_impact": stock, "match_status": status, "confidence": conf})
    # luu profile neu yeu cau (tu hoc) — dung chung cho MOI scope
    if params.get("save_profile"):
        _, all_rows = _load_sheet_rows(file_bytes, sheet=sheet, max_row=header_row + 1)
        sig = _file_signature(all_rows[header_row]) if header_row < len(all_rows) else None
        if sig:
            conn.execute("""INSERT INTO import_profile(ten_profile, scope, doi_tac, file_signature,
                sheet_name, header_row, data_start_row, col_map)
                VALUES(?,?,?,?,?,?,?,?)
                ON CONFLICT(file_signature) DO UPDATE SET col_map=excluded.col_map,
                scope=excluded.scope, sheet_name=excluded.sheet_name, header_row=excluded.header_row,
                data_start_row=excluded.data_start_row, ten_profile=excluded.ten_profile""",
                (params.get("ten_profile") or "Profile " + (target.get("supplier_name") or ""),
                 scope, target.get("supplier_name") or target.get("customer_id"),
                 sig, sheet, header_row, data_start, json.dumps(col_map)))
            conn.commit()
    tok = "flex_" + secrets.token_urlsafe(12)
    _TOKENS[tok] = {"lines": lines, "scope": scope, "target": target, "het_han": time.time() + 600}
    so_pending = sum(1 for x in lines if x["match_status"] in ("pending", "unmatched"))
    return {"ok": True, "confirm_token": tok, "lines": lines[:500], "so_dong": len(lines),
            "so_pending": so_pending, "scope": scope}


def import_flex_commit(conn, sess, token, overrides=None):
    """Ghi theo scope: bao_gia_ncc -> item_cost_history(source_type='bao_gia_ncc');
    hoa_don_dau_vao -> 1 hoa_don(mua_vao)+dong+ledger; moi_thau_khach -> import_flex_line tam."""
    from api_write import audit
    entry = _TOKENS.pop(token, None)
    if not entry:
        raise ValueError("Token het han hoac da dung — lam lai buoc map.")
    if entry["het_han"] < time.time():
        raise ValueError("Token het han (10 phut).")
    lines = entry["lines"]
    scope = entry["scope"]
    target = entry["target"]
    ov = {o.get("ten_hang"): o for o in (overrides or [])}
    for ln in lines:
        o = ov.get(ln["ten_hang"])
        if o:
            # .get(...) an toan cho scope bbnt_cu/pxk_cu (khong co item_key/cost_type)
            ln["item_key"] = o.get("item_key") or ln.get("item_key")
            ln["cost_type"] = o.get("cost_type") or ln.get("cost_type")
            if o.get("stock_impact") is not None:
                ln["stock_impact"] = o["stock_impact"]
            ln["match_status"] = "confirmed"

    st = {"scope": scope, "so_dong": len(lines), "cost_rows": 0, "hoa_don": 0, "flex_line": 0}
    today = date.today().isoformat()
    if scope == "bao_gia_ncc":
        ncc = target.get("supplier_name") or "NCC"
        for ln in lines:
            if ln["cost_type"] in ("thiet_bi", "vat_tu") and ln["match_status"] in ("auto", "confirmed"):
                conn.execute("""INSERT INTO item_cost_history(item_key, item_name, item_group, uom,
                    supplier_name, purchase_date, quantity, unit_cost, vat_rate, cost_with_vat, source_type)
                    VALUES(?,?,?,?,?,?,?,?,?,?,?)""",
                    (ln["item_key"], (ln["ten_hang"] + " " + ln["model"]).strip(), ln["cost_type"],
                     ln["dvt"], ncc, today, ln["so_luong"], ln["don_gia"], ln["thue_suat"],
                     ln["don_gia"] * (1 + ln["thue_suat"] / 100.0), "bao_gia_ncc"))
                st["cost_rows"] += 1
    elif scope == "hoa_don_dau_vao":
        ncc = target.get("supplier_name") or "NCC"
        mst = target.get("mst") or ""
        ma_hd = target.get("ma_hd") or ("FLEX-" + _stamp())
        tong_tt = sum(l["don_gia"] * l["so_luong"] for l in lines)
        tong_thue = sum(l["don_gia"] * l["so_luong"] * l["thue_suat"] / 100.0 for l in lines)
        conn.execute("""INSERT INTO hoa_don(ma_hd, ngay, customer_id, mst, ten_don_vi,
            tong_truoc_thue, tong_thue, tong_cong, chieu, nguon_file)
            VALUES(?,?,?,?,?,?,?,?,?,?)""",
            (ma_hd, today, None, mst, ncc, tong_tt, tong_thue, tong_tt + tong_thue,
             "mua_vao", "import_flex"))
        hd_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        st["hoa_don"] = 1
        for ln in lines:
            tt = ln["don_gia"] * ln["so_luong"]
            conn.execute("""INSERT INTO hoa_don_dong(hoa_don_id, ten_hang_hoa, dvt, so_luong,
                don_gia, thanh_tien, thue_suat, tien_thue, cost_type, stock_impact, item_key,
                match_confidence, match_status) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (hd_id, (ln["ten_hang"] + " " + ln["model"]).strip(), ln["dvt"], ln["so_luong"],
                 ln["don_gia"], tt, ln["thue_suat"], tt * ln["thue_suat"] / 100.0,
                 ln["cost_type"], ln["stock_impact"], ln["item_key"], ln["confidence"],
                 ln["match_status"]))
            dong_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
            if ln["stock_impact"] and ln["match_status"] in ("auto", "confirmed"):
                grp = {"ncc": ncc, "mst": mst, "ngay": today, "ma_hd": ma_hd}
                d = {"ten_hang_hoa": ln["ten_hang"], "dvt": ln["dvt"], "so_luong": ln["so_luong"],
                     "don_gia": ln["don_gia"], "thanh_tien": tt, "thue_suat": ln["thue_suat"],
                     "tien_thue": tt * ln["thue_suat"] / 100.0, "cost_type": ln["cost_type"]}
                conn.execute("""INSERT INTO item_cost_history(item_key, item_name, item_group, uom,
                    supplier_name, supplier_mst, hoa_don_id, hoa_don_dong_id, purchase_date, quantity,
                    unit_cost, vat_rate, cost_with_vat, source_type)
                    VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (ln["item_key"], d["ten_hang_hoa"], d["cost_type"], d["dvt"], ncc, mst, hd_id,
                     dong_id, today, d["so_luong"], d["don_gia"], d["thue_suat"],
                     d["thanh_tien"] + d["tien_thue"], "mua_vao"))
                conn.execute("""INSERT INTO stock_ledger(item_key, item_name, movement_type,
                    source_type, source_id, source_line_id, movement_date, qty_in, unit_cost, amount, note)
                    VALUES(?,?,?,?,?,?,?,?,?,?,?)""",
                    (ln["item_key"], d["ten_hang_hoa"], "nhap_mua", "mua_vao", hd_id, dong_id,
                     today, d["so_luong"], d["don_gia"], d["thanh_tien"], "Flex " + ma_hd))
                HM._cap_nhat_gia_von_catalog(conn, ln["item_key"], d, grp)
                st["cost_rows"] += 1
    elif scope == "moi_thau_khach":  # luu tam import_flex_line (cho tao bao gia) — GIU NGUYEN
        existing_cols = {r[1] for r in conn.execute("PRAGMA table_info(import_flex_line)").fetchall()}
        if "thue_suat" not in existing_cols:
            conn.execute("ALTER TABLE import_flex_line ADD COLUMN thue_suat REAL")
        if "thanh_tien" not in existing_cols:
            conn.execute("ALTER TABLE import_flex_line ADD COLUMN thanh_tien REAL")
        batch = "batch_" + _stamp()
        for ln in lines:
            conn.execute("""INSERT INTO import_flex_line(batch, scope, customer_id, project_id,
                supplier_name, ten_hang, model, dvt, so_luong, don_gia, thanh_tien, thue_suat,
                item_key, cost_type, match_status)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (batch, scope, target.get("customer_id"), target.get("project_id"),
                 target.get("supplier_name"), ln["ten_hang"], ln["model"], ln["dvt"],
                 ln["so_luong"], ln["don_gia"], ln.get("thanh_tien"), ln.get("thue_suat"), ln["item_key"],
                 ln["cost_type"], ln["match_status"]))
            st["flex_line"] += 1
        st["batch"] = batch
    else:
        # WO-29 Phase 1: bbnt_cu/pxk_cu KHONG di qua commit chung (hinh dang dong khac hang
        # hoa, INSERT vao import_flex_line se mat field rieng nhu hang_muc/kl_hop_dong) —
        # dung bridge rieng write/tao_bbnt_tu_list hoac write/tao_pxk_tu_list.
        raise ValueError("Scope '%s' dùng bridge riêng (tao_bbnt_tu_list/tao_pxk_tu_list), "
                         "không commit qua import_flex_commit." % scope)
    audit(conn, sess, "import_flex", "import_flex", scope, "Flex commit %s: %s" % (scope, st))
    conn.commit()
    return st


def _stamp():
    return datetime.now().strftime("%Y%m%d%H%M%S")


def lay_lines_token(token):
    e = _TOKENS.get(token)
    return e["lines"] if e else None
