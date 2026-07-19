# -*- coding: utf-8 -*-
"""Two-phase project-profile import for App 8777.

The preview phase is pure/read-only.  The commit phase writes one audited SQLite
transaction and preserves previous quotations/BOQ imports as history.  Original
business files are only indexed when a local path is supplied; uploaded bytes are
saved as a new, non-overwriting source file below D:\\2025 or D:\\2026.
"""
from __future__ import annotations

import base64
import hashlib
import json
import os
import re
import secrets
import sqlite3
import threading
import time
import unicodedata
from collections import Counter
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

import openpyxl

import contract_pl01_compare as CONTRACT_PL01
import boq_normalizer as QUOTE_PARSER


ALLOWED_SOURCE_ROOTS = (r"D:\2025", r"D:\2026")
# Official quotation/profile imports expose price and total values.  KTT gets a
# separate quantity/stage assignment projection and must never receive money.
PROFILE_ROLES = ("Giam doc", "Ke toan", "Quan tri he thong")
PREVIEW_TTL_SECONDS = 30 * 60
MAX_FILE_BYTES = 25 * 1024 * 1024
PROJECT_NAME_DEFAULT = (
    "Cung cap va lap dat he thong dieu hoa khong khi va thong gio - "
    "Khoi nha xu ly ve sinh tau bay"
)
DEFAULT_TEMPLATE_PROFILE = "INSTALLATION_STANDARD"


class ProfileImportError(ValueError):
    pass


_PREVIEWS = {}
_PREVIEW_LOCK = threading.Lock()


def _require_role(sess):
    if (sess or {}).get("role") not in PROFILE_ROLES:
        raise ProfileImportError("Vai tro hien tai khong co quyen import ho so cong trinh.")


def _template_profile_from_input(data):
    profile = (data.get("template_profile") or DEFAULT_TEMPLATE_PROFILE).strip()
    try:
        import docgen
        profiles = set(docgen.ct_document_profiles())
    except Exception as exc:
        raise ProfileImportError("Khong nap duoc danh muc profile V3.1: %s" % exc)
    if profile not in profiles:
        raise ProfileImportError("Loai ho so V3.1 khong hop le: %s." % profile)
    return profile


def _norm(value):
    text = unicodedata.normalize("NFD", str(value or ""))
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = text.replace("Đ", "D").replace("đ", "d").casefold()
    return re.sub(r"\s+", " ", text).strip()


def _dec(value):
    if value in (None, ""):
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError, TypeError):
        return None


def _num(value):
    parsed = _dec(value)
    return float(parsed) if parsed is not None else None


def _raw(payload):
    if not isinstance(payload, dict):
        return None if payload is None else str(payload)
    value = payload.get("raw_value")
    if value is None:
        return None
    return str(value)


def _sha256_bytes(payload):
    return hashlib.sha256(payload).hexdigest()


def _is_under(path, roots=ALLOWED_SOURCE_ROOTS):
    # Resolve junctions/symlinks before applying the allow-root check.
    target = os.path.normcase(os.path.realpath(os.path.abspath(path)))
    for root in roots:
        root_abs = os.path.normcase(os.path.realpath(os.path.abspath(root)))
        try:
            if os.path.commonpath((target, root_abs)) == root_abs:
                return True
        except ValueError:
            continue
    return False


def _decode_b64(value):
    text = str(value or "")
    if "," in text and text.split(",", 1)[0].startswith("data:"):
        text = text.split(",", 1)[1]
    try:
        return base64.b64decode(text, validate=True)
    except Exception as exc:
        raise ProfileImportError("Du lieu file base64 khong hop le.") from exc


def _file_input(data, key, allowed_exts, required=False):
    path = (data.get(key + "_path") or "").strip()
    filename = (data.get(key + "_filename") or "").strip()
    encoded = data.get(key + "_b64")
    payload = None
    if path:
        # Pin the resolved target used for hashing/indexing.  Parsing below is
        # from the captured bytes, so a later path swap cannot change preview.
        path = os.path.realpath(os.path.abspath(path))
        if not _is_under(path):
            raise ProfileImportError("File %s phai nam duoi D:\\2025 hoac D:\\2026." % key)
        if not os.path.isfile(path):
            raise ProfileImportError("Khong tim thay file %s." % key)
        filename = os.path.basename(path)
        size = os.path.getsize(path)
        if size > MAX_FILE_BYTES:
            raise ProfileImportError("File %s vuot gioi han %d MB." % (key, MAX_FILE_BYTES // 1024 // 1024))
        with open(path, "rb") as handle:
            payload = handle.read()
    elif encoded:
        payload = _decode_b64(encoded)
        if not filename:
            raise ProfileImportError("Thieu ten file %s upload." % key)
        if len(payload) > MAX_FILE_BYTES:
            raise ProfileImportError("File %s vuot gioi han %d MB." % (key, MAX_FILE_BYTES // 1024 // 1024))
    elif required:
        raise ProfileImportError("Thieu file %s." % key)
    else:
        return None
    ext = os.path.splitext(filename)[1].lower()
    if ext not in allowed_exts:
        raise ProfileImportError("Dinh dang file %s khong duoc ho tro: %s" % (key, ext or "(rong)"))
    return {
        "key": key,
        "path": path or None,
        "filename": filename,
        "ext": ext,
        "size": len(payload),
        "sha256": _sha256_bytes(payload),
        "bytes": payload,
    }


def _excel_value(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return "" if value is None else str(value).strip()


def _preview_personnel(file_info):
    if not file_info:
        return {"rows": [], "warnings": [], "sheet_name": None}
    try:
        wb = openpyxl.load_workbook(
            filename=__import__("io").BytesIO(file_info["bytes"]),
            data_only=True, read_only=True, keep_links=False,
        )
    except Exception as exc:
        raise ProfileImportError("Khong doc duoc danh sach nhan su XLSX: %s" % exc) from exc
    try:
        ws = wb[wb.sheetnames[0]]
        header_row = None
        headers = []
        for row_no in range(1, min(ws.max_row, 15) + 1):
            vals = [_norm(ws.cell(row_no, col).value) for col in range(1, ws.max_column + 1)]
            if any(("ho ten" in value or "ho va ten" in value) for value in vals) and \
                    any(("cccd" in value or "cmnd" in value) for value in vals):
                header_row, headers = row_no, vals
                break
        if not header_row:
            raise ProfileImportError("Khong nhan dien duoc dong tieu de danh sach nhan su.")

        def positions(*needles):
            return [idx + 1 for idx, value in enumerate(headers)
                    if any(needle in value for needle in needles)]

        name_cols = positions("ho ten", "ho va ten")
        cccd_cols = positions("cccd", "cmnd")
        role_cols = positions("chuc vu", "vai tro")
        if not name_cols or not cccd_cols:
            raise ProfileImportError("Danh sach nhan su thieu cot Ho ten/CCCD.")
        col = {
            "stt": (positions("stt") or [1])[0],
            "mst": (positions("mst", "ma so thue") or [None])[0],
            "name": name_cols[0],
            "birth": (positions("nam sinh", "ngay sinh") or [None])[0],
            "cccd": cccd_cols[0],
            "subcontractor": (positions("nha thau phu", "don vi") or [None])[0],
            "site_role": role_cols[0] if role_cols else None,
            "project_role": role_cols[-1] if role_cols else None,
            "expiry": (positions("het han") or [None])[0],
            "issue": (positions("ngay cap") or [None])[0],
            "locked": (positions("khoa the", "khoa") or [None])[0],
        }
        rows, warnings = [], []
        for row_no in range(header_row + 1, ws.max_row + 1):
            name = _excel_value(ws.cell(row_no, col["name"]).value)
            cccd = _excel_value(ws.cell(row_no, col["cccd"]).value)
            if not name and not cccd:
                continue
            if not name:
                warnings.append({"code": "PERSONNEL_NAME_MISSING", "row": row_no})
                continue
            if cccd and not re.fullmatch(r"\d{9,12}", cccd):
                warnings.append({"code": "PERSONNEL_ID_FORMAT", "row": row_no})
            value = lambda key: (_excel_value(ws.cell(row_no, col[key]).value)
                                 if col.get(key) else "")
            rows.append({
                "source_row": row_no,
                "source_stt": value("stt"),
                "source_mst": value("mst"),
                "name": name,
                "birth": value("birth"),
                "cccd": cccd,
                "subcontractor": value("subcontractor"),
                "site_role": value("site_role"),
                "project_role": value("project_role"),
                "card_expiry": value("expiry"),
                "card_issue_date": value("issue"),
                "card_locked": value("locked"),
            })
        return {"rows": rows, "warnings": warnings, "sheet_name": ws.title}
    finally:
        wb.close()


def _find_project(conn, project_id, project_name, customer_id):
    if project_id:
        row = conn.execute("SELECT * FROM project WHERE id=?", (project_id,)).fetchone()
        if not row:
            raise ProfileImportError("Cong trinh duoc chon khong ton tai.")
        if customer_id and int(customer_id) != int(row["customer_id"]):
            raise ProfileImportError("Cong trinh khong thuoc khach hang duoc chon.")
        return row
    wanted = _norm(project_name)
    matches = [row for row in conn.execute("SELECT * FROM project ORDER BY id").fetchall()
               if _norm(row["project_name"]) == wanted and
               (not customer_id or int(row["customer_id"]) == int(customer_id))]
    if len(matches) > 1:
        raise ProfileImportError("Co nhieu cong trinh trung ten; hay chon project_id cu the.")
    return matches[0] if matches else None


def _existing_contract(conn, project_id, explicit_id=None):
    if explicit_id:
        row = conn.execute("SELECT * FROM hop_dong_ct WHERE id=?", (explicit_id,)).fetchone()
        if not row:
            raise ProfileImportError("Hop dong duoc chon khong ton tai.")
        q = conn.execute("SELECT project_id FROM quotation WHERE id=?", (row["quotation_id"],)).fetchone()
        if project_id and q and int(q["project_id"] or 0) != int(project_id):
            raise ProfileImportError("Hop dong khong thuoc cong trinh duoc chon.")
        return row
    if not project_id:
        return None
    return conn.execute("""SELECT h.* FROM hop_dong_ct h
        JOIN quotation q ON q.id=h.quotation_id
        WHERE q.project_id=? ORDER BY h.id DESC LIMIT 1""", (project_id,)).fetchone()


def _cleanup_previews(now=None):
    now = now or time.time()
    with _PREVIEW_LOCK:
        expired = [token for token, state in _PREVIEWS.items()
                   if state["created_at"] + PREVIEW_TTL_SECONDS < now]
        for token in expired:
            _PREVIEWS.pop(token, None)


def preview_project_profile(conn, sess, data):
    """Read/parse only.  Returns a short-lived confirmation token."""
    _require_role(sess)
    _cleanup_previews()
    project_name = (data.get("project_name") or PROJECT_NAME_DEFAULT).strip()
    if not project_name:
        raise ProfileImportError("Thieu ten cong trinh.")
    template_profile = _template_profile_from_input(data)
    quote_file = _file_input(data, "quote", {".xls", ".xlsx", ".xlsm", ".xlsb"}, required=True)
    contract_file = _file_input(data, "contract", {".doc", ".docx", ".pdf"})
    personnel_file = _file_input(data, "personnel", {".xlsx", ".xlsm"})
    try:
        parsed_quote = QUOTE_PARSER.normalize_workbook(
            quote_file["bytes"], quote_file["filename"], data.get("quote_sheet_name"))
    except QUOTE_PARSER.BoqNormalizationError as exc:
        raise ProfileImportError(str(exc)) from exc
    parsed_personnel = _preview_personnel(personnel_file)
    project = _find_project(conn, data.get("project_id"), project_name, data.get("customer_id"))
    customer_id = int(project["customer_id"]) if project else int(data.get("customer_id") or 0)
    if not customer_id or not conn.execute("SELECT 1 FROM customer WHERE id=?", (customer_id,)).fetchone():
        raise ProfileImportError("Hay chon khach hang hop le cho cong trinh.")
    contract = _existing_contract(conn, project["id"] if project else None, data.get("contract_id"))
    warnings = list(parsed_quote.get("warnings") or []) + list(parsed_personnel["warnings"])
    contract_comparison = None
    if contract_file and contract_file["ext"] == ".doc":
        # Legacy Word is inspected from its captured bytes through a private,
        # macro-disabled/read-only shadow instance.  The comparison result is
        # preview-only and contains codes/row numbers/counts, never cell text,
        # personnel data, prices, or amounts.
        contract_comparison = CONTRACT_PL01.compare_legacy_doc_pl01(
            contract_file["bytes"], parsed_quote,
        )
        warnings.extend(contract_comparison.get("discrepancies") or [])
    if contract_file and contract_file["ext"] == ".doc" and not contract:
        warnings.append({"code": "LEGACY_DOC_INDEX_ONLY"})
    stage_names = [stage["name"] for stage in parsed_quote["stages"]]
    # The selected project/contract are part of import identity.  The same
    # source files may legitimately be linked elsewhere and are not a replay.
    bundle_parts = [
        "quote:%s" % quote_file["sha256"],
        "contract_file:%s" % (contract_file["sha256"] if contract_file else ""),
        "personnel:%s" % (personnel_file["sha256"] if personnel_file else ""),
        # Stable across the create -> subsequent-preview transition.  The SQL
        # idempotency lookup is also scoped by project_id, so a normalized name
        # is sufficient here and cannot collide across projects.
        "project_name:%s" % _norm(project_name),
        "customer:%s" % customer_id,
        "contract:%s" % (contract["id"] if contract else "none"),
        "template_profile:%s" % template_profile,
    ]
    bundle_sha = _sha256_bytes("|".join(bundle_parts).encode("utf-8"))
    state = {
        "created_at": time.time(), "project_id": project["id"] if project else None,
        "project_name": project_name, "customer_id": customer_id,
        "template_profile": template_profile,
        "contract_id": contract["id"] if contract else None,
        "quote_file": quote_file, "contract_file": contract_file,
        "personnel_file": personnel_file, "quote": parsed_quote,
        "personnel": parsed_personnel, "contract_comparison": contract_comparison,
        "warning_metadata": warnings, "bundle_sha256": bundle_sha,
        "auto_generate_templates": data.get("auto_generate_templates", True) is not False,
    }
    token = secrets.token_urlsafe(32)
    with _PREVIEW_LOCK:
        _PREVIEWS[token] = state
    warning_counts = Counter(item.get("code", "OTHER") for item in warnings)
    return {
        "ok": True,
        "confirm_token": token,
        "expires_in_seconds": PREVIEW_TTL_SECONDS,
        "project": {
            "id": project["id"] if project else None,
            "code": project["code"] if project else None,
            "project_name": project_name,
            "template_profile": template_profile,
            "mode": "update_in_place" if project else "create",
        },
        "quote": {
            "sheet_name": parsed_quote["source"]["sheet_name"],
            "allocation_sheet_name": parsed_quote["source"].get("allocation_sheet_name"),
            "document_kind": parsed_quote.get("document_kind"),
            "sha256_prefix": quote_file["sha256"][:12],
            "counts": parsed_quote["counts"],
            "stages": stage_names,
            "invariants": parsed_quote["invariants"],
            "candidate_sheets": parsed_quote.get("candidate_sheets") or [],
            "normalization_audit": parsed_quote.get("normalization_audit") or {},
            "can_commit_official": bool(
                parsed_quote.get("document_kind") == "QUOTATION" and
                (parsed_quote.get("normalization_audit") or {}).get("ready_for_official_commit")),
        },
        "personnel": {
            "count": len(parsed_personnel["rows"]),
            "project_roles": dict(Counter(row["project_role"] or "Chua ghi" for row in parsed_personnel["rows"])),
        },
        "contract": {
            "file_supplied": bool(contract_file),
            "linked_existing_contract": bool(contract),
            "legacy_doc_index_only": bool(contract_file and contract_file["ext"] == ".doc" and not contract),
            "pl01_comparison": contract_comparison,
        },
        "warning_counts": dict(warning_counts),
        "planned_actions": [
            "preserve_project_contract_workflow_drafts" if project else "create_project",
            "create_official_quotation_revision",
            "replace_unused_predicted_budget_only",
            "import_exact_stage_boq",
            "upsert_project_personnel",
            "index_original_source_files",
            "wire_template_registry_and_personnel_documents",
        ],
    }


def _verify_files_unchanged(state):
    for key in ("quote_file", "contract_file", "personnel_file"):
        info = state.get(key)
        if not info or not info.get("path"):
            continue
        if not _is_under(info["path"]):
            raise ProfileImportError("Duong dan file %s khong con nam trong vung nguon cho phep." % key)
        try:
            with open(info["path"], "rb") as handle:
                digest = _sha256_bytes(handle.read())
        except OSError as exc:
            raise ProfileImportError("Khong doc lai duoc file %s truoc khi commit." % key) from exc
        if digest != info["sha256"]:
            raise ProfileImportError("File %s da thay doi sau preview; hay preview lai." % key)


def _backup_database(conn):
    """Create a consistent SQLite online backup immediately before profile writes."""
    row = conn.execute("PRAGMA database_list").fetchone()
    path = row[2] if row and len(row) > 2 else None
    if not path or not os.path.isfile(path):
        return None
    backup_path = "%s.bak_before_project_profile_%s" % (
        path, datetime.now().strftime("%Y%m%d_%H%M%S_%f"))
    target = sqlite3.connect(backup_path)
    try:
        conn.backup(target)
    finally:
        target.close()
    return backup_path


def _safe_filename(name):
    value = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", os.path.basename(name or "file"))
    return value.strip(" .") or "file"


def _customer_root(conn, customer_id, state):
    for info in (state.get("quote_file"), state.get("contract_file"), state.get("personnel_file")):
        path = info.get("path") if info else None
        if path and _is_under(path):
            parts = os.path.abspath(path).split(os.sep)
            if len(parts) >= 3 and parts[1].isdigit():
                return os.path.join(parts[0] + os.sep, parts[1], parts[2])
    customer = conn.execute("SELECT customer_name, duong_dan_folder FROM customer WHERE id=?",
                            (customer_id,)).fetchone()
    if customer and customer["duong_dan_folder"] and _is_under(customer["duong_dan_folder"]):
        return customer["duong_dan_folder"]
    row = conn.execute("""SELECT nam_nguon, khach_folder FROM source_document
        WHERE customer_id=? AND nam_nguon IN ('2025','2026') AND khach_folder IS NOT NULL
        ORDER BY id DESC LIMIT 1""", (customer_id,)).fetchone()
    if row:
        return os.path.join("D:\\", row["nam_nguon"], row["khach_folder"])
    name = _safe_filename(customer["customer_name"] if customer else "Khach hang")
    return os.path.join(r"D:\2026", "%s %s" % (name, customer_id))


def _materialize_upload(conn, customer_id, state, info, subfolder):
    if not info or info.get("path"):
        return None
    root = _customer_root(conn, customer_id, state)
    if not _is_under(root):
        raise ProfileImportError("Khong xac dinh duoc folder nguon an toan cho file upload.")
    dest_dir = os.path.join(root, subfolder)
    created_dir = not os.path.isdir(dest_dir)
    os.makedirs(dest_dir, exist_ok=True)
    filename = _safe_filename(info["filename"])
    path = os.path.join(dest_dir, filename)
    if os.path.exists(path):
        with open(path, "rb") as handle:
            existing_sha = _sha256_bytes(handle.read())
        if existing_sha != info["sha256"]:
            stem, ext = os.path.splitext(filename)
            path = os.path.join(dest_dir, "%s_%s_%s%s" %
                                (stem, datetime.now().strftime("%Y%m%d_%H%M%S"), info["sha256"][:8], ext))
            counter = 2
            while os.path.exists(path):
                path = os.path.join(dest_dir, "%s_%s_%s_%d%s" %
                                    (stem, datetime.now().strftime("%Y%m%d_%H%M%S"),
                                     info["sha256"][:8], counter, ext))
                counter += 1
    created_file = not os.path.exists(path)
    if created_file:
        # Exclusive creation prevents an accidental overwrite if imports race.
        with open(path, "xb") as handle:
            handle.write(info["bytes"])
    info["path"] = path
    return {"path": path if created_file else None,
            "dir": dest_dir if created_dir else None, "info": info}


def _cleanup_materialized_uploads(created):
    """Best-effort compensation for filesystem writes after SQL rollback."""
    for item in reversed(created):
        path = item.get("path")
        if path and _is_under(path):
            try:
                if os.path.isfile(path):
                    os.remove(path)
            except OSError:
                pass
        info = item.get("info")
        if info is not None and path and info.get("path") == path:
            info["path"] = None
        folder = item.get("dir")
        if folder and _is_under(folder):
            try:
                os.rmdir(folder)  # succeeds only if this import left it empty
            except OSError:
                pass


def _index_source(conn, customer_id, project_id, info, profile_role, doc_type):
    if not info:
        return None
    path = os.path.realpath(os.path.abspath(info["path"]))
    if not _is_under(path):
        raise ProfileImportError("Khong index file nam ngoai vung nguon cho phep.")
    stat = os.stat(path)
    parts = path.split(os.sep)
    year = parts[1] if len(parts) > 1 and parts[1].isdigit() else None
    root = os.path.join(parts[0] + os.sep, year) if year else os.path.dirname(path)
    khach_folder = parts[2] if year and len(parts) > 2 else os.path.basename(os.path.dirname(path))
    rel_path = os.path.relpath(path, root)
    existing = conn.execute("SELECT id FROM source_document WHERE lower(abs_path)=lower(?) ORDER BY id LIMIT 1",
                            (path,)).fetchone()
    values = (customer_id, project_id, profile_role, khach_folder, doc_type,
              os.path.basename(path), rel_path, path, info["ext"], stat.st_size,
              info["sha256"], datetime.fromtimestamp(stat.st_mtime).isoformat(timespec="seconds"), year)
    if existing:
        conn.execute("""UPDATE source_document SET customer_id=?, project_id=?, profile_role=?,
            khach_folder=?, doc_type=?, file_name=?, rel_path=?, abs_path=?, ext=?, size_bytes=?,
            source_sha256=?, mtime=?, nam_nguon=?, scanned_at=datetime('now') WHERE id=?""",
                     values + (existing["id"],))
        return existing["id"]
    conn.execute("""INSERT INTO source_document(customer_id, project_id, profile_role,
        khach_folder, doc_type, file_name, rel_path, abs_path, ext, size_bytes,
        source_sha256, mtime, nam_nguon, scanned_at)
        VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,datetime('now'))""", values)
    return conn.execute("SELECT last_insert_rowid()").fetchone()[0]


def _next_code(conn, table, prefix):
    year = date.today().year
    row = conn.execute("SELECT MAX(CAST(SUBSTR(code,-4) AS INTEGER)) FROM %s WHERE code LIKE ?" % table,
                       ("%s-%d-%%" % (prefix, year),)).fetchone()
    return "%s-%d-%04d" % (prefix, year, (row[0] or 0) + 1)


def _new_quote_code(conn, current):
    if not current:
        return _next_code(conn, "quotation", "BG")
    base = re.sub(r"-V\d+$", "", current["code"])
    version = 2
    while conn.execute("SELECT 1 FROM quotation WHERE code=?", ("%s-V%d" % (base, version),)).fetchone():
        version += 1
    return "%s-V%d" % (base, version)


def _vat_for_line(line):
    # In the official workbook column P is the pre-tax line amount.  The tax
    # rate is an explicit source note such as "VAT 8%" or "VAT 10%".
    note = str(line.get("note_raw") or "")
    match = re.search(r"(?<!\d)(8|10)\s*%", note, flags=re.IGNORECASE)
    if match and ("vat" in note.casefold() or "thu" in _norm(note)):
        return Decimal(match.group(1))
    qty = _dec((line.get("contract_quantity") or {}).get("value"))
    price = _dec((line.get("unit_price") or {}).get("value"))
    amount = _dec((line.get("amount") or {}).get("value"))
    if qty is None or price is None or amount is None or qty * price == 0:
        return Decimal(0)
    rate = (amount / (qty * price) - Decimal(1)) * Decimal(100)
    for expected in (Decimal(8), Decimal(10)):
        if abs(rate - expected) <= Decimal("0.05"):
            return expected
    return rate.quantize(Decimal("0.0001")) if Decimal(-100) <= rate <= Decimal(100) else Decimal(0)


def _parse_quote_footer_totals(totals):
    """Read VAT / grand from official footer labels (not per-line notes).

    Typical Long Thanh quote footer:
      TỔNG CỘNG (THUẾ 10%) / THUẾ VAT 10% / TỔNG CỘNG (1)
      TỔNG CỘNG (THUẾ 8%)  / THUẾ VAT 8%  / TỔNG CỘNG (2)
      THÀNH TIỀN (TC1 + TC2)
    Detail rows are pre-tax; line notes are often empty, so VAT must come here.
    """
    vat8 = None
    vat10 = None
    before_parts = []
    grand = None
    for total in totals or []:
        label = str(total.get("text_raw") or total.get("stt_raw") or "")
        label_n = _norm(label)
        amount = _dec((total.get("amount") or {}).get("value"))
        if amount is None:
            continue
        # Explicit VAT lines: "THUẾ VAT 8%", "THUẾ VAT 10%"
        if "thue vat" in label_n or re.search(r"\bvat\s*(8|10)\s*%", label_n):
            match = re.search(r"(?<!\d)(8|10)\s*%", label)
            if match:
                if match.group(1) == "8":
                    vat8 = amount
                else:
                    vat10 = amount
            continue
        if "thanh tien" in label_n:
            grand = amount
            continue
        # Pre-tax group bases: "TỔNG CỘNG (THUẾ 8%)" / "TỔNG CỘNG (THUẾ 10%)"
        # (not "TỔNG CỘNG (1)" which is after-tax).
        if "tong cong" in label_n and re.search(r"thue\s*(8|10)\s*%", label_n):
            before_parts.append(amount)
            continue
    if grand is None:
        for total in reversed(totals or []):
            amount = _dec((total.get("amount") or {}).get("value"))
            if amount is not None:
                grand = amount
                break
    tax = None
    if vat8 is not None or vat10 is not None:
        tax = (vat8 or Decimal(0)) + (vat10 or Decimal(0))
    return {
        "vat8": vat8,
        "vat10": vat10,
        "tax": tax,
        "before": sum(before_parts, Decimal(0)) if before_parts else None,
        "grand": grand,
    }


def _quote_totals(parsed):
    before = Decimal(0)
    line_tax = Decimal(0)
    line_vat8 = Decimal(0)
    line_vat10 = Decimal(0)
    for line in parsed["lines"]:
        if line["kind"] != "detail":
            continue
        amount = _dec(line["amount"].get("value")) or Decimal(0)
        rate = _vat_for_line(line)
        # Line notes may carry VAT; amounts themselves are usually pre-tax.
        lt = amount * rate / Decimal(100)
        before += amount
        line_tax += lt
        if rate == 8:
            line_vat8 += lt
        elif rate == 10:
            line_vat10 += lt

    footer = _parse_quote_footer_totals(parsed.get("totals") or [])
    # Prefer footer VAT when the workbook groups tax only in the footer
    # (common for official quotes: note_raw empty → line rate = 0).
    if footer.get("tax") is not None:
        vat8 = footer["vat8"] or Decimal(0)
        vat10 = footer["vat10"] or Decimal(0)
        tax = footer["tax"]
    else:
        vat8, vat10, tax = line_vat8, line_vat10, line_tax

    if footer.get("grand") is not None:
        grand = footer["grand"]
    else:
        grand = before + tax

    # Detail pre-tax sum must match footer group bases when present.
    if footer.get("before") is not None and abs(footer["before"] - before) > Decimal("1"):
        raise ProfileImportError(
            "Tong hang chua thue khong khop nhom TONG CONG (THUE x%%) o cuoi bao gia.")
    expected = before + tax
    if abs(expected - grand) > Decimal("1"):
        raise ProfileImportError(
            "Tong tien tinh tu chi tiet khong khop tong cuoi trong bao gia chinh thuc.")
    return {"before": before, "tax": tax, "grand": grand, "vat8": vat8, "vat10": vat10}


def _current_quote(conn, project_id):
    return conn.execute("""SELECT q.* FROM quotation q WHERE q.project_id=? AND q.status<>'Huy'
        AND NOT EXISTS (SELECT 1 FROM quotation child WHERE child.amended_from=q.id)
        ORDER BY q.id DESC LIMIT 1""", (project_id,)).fetchone()


def _insert_official_quote(conn, project_id, customer_id, current, state):
    parsed = state["quote"]
    totals = _quote_totals(parsed)
    cols = [row["name"] for row in conn.execute("PRAGMA table_info(quotation)").fetchall()
            if row["name"] not in ("id", "created_at")]
    base = dict(current) if current else {}
    values = dict(base)
    values.update({
        "code": _new_quote_code(conn, current), "customer_id": customer_id,
        "project_id": project_id, "grand_total": float(totals["grand"]),
        "tong_truoc_thue": float(totals["before"]), "tien_thue": float(totals["tax"]),
        "vat_8": float(totals["vat8"]), "vat_10": float(totals["vat10"]),
        "status": "Da_duyet", "amended_from": current["id"] if current else None,
        "ngay_lap": date.today().isoformat(), "trang_thai_doi_chieu": "xong",
        "source_file_name": state["quote_file"]["filename"],
        "source_sha256": state["quote_file"]["sha256"], "is_official": 1,
        "imported_at": datetime.now().isoformat(timespec="seconds"),
    })
    insert_cols = [col for col in cols if col in values]
    conn.execute("INSERT INTO quotation(%s) VALUES(%s)" %
                 (",".join(insert_cols), ",".join("?" for _ in insert_cols)),
                 [values[col] for col in insert_cols])
    quote_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    item_columns = {row["name"] for row in conn.execute("PRAGMA table_info(quotation_item)").fetchall()}
    item_ids = {}
    for line in parsed["lines"]:
        detail = line["kind"] == "detail"
        qty = _num(line["contract_quantity"].get("value")) if detail else None
        price = _num(line["unit_price"].get("value")) if detail else 0
        amount = _num(line["amount"].get("value")) if detail else 0
        vat = float(_vat_for_line(line)) if detail else None
        allocations = [cell["stage_name_raw"] for cell in line.get("stage_quantities", [])]
        if detail and not allocations:
            allocations = [QUOTE_PARSER.UNALLOCATED_STAGE_NAME]
        record = {
            "quotation_id": quote_id, "stt": line["source_order"],
            "hang_muc": str(line.get("text_raw") or ""),
            "khoi_luong": line["contract_quantity"].get("value") if detail else None,
            "don_gia": price or 0, "thanh_tien": amount or 0,
            "trang_thai": "Da_import_chinh_thuc", "thue_suat": vat,
            "tien_thue": (float((_dec(line["amount"].get("value")) or Decimal(0))
                                * _vat_for_line(line) / Decimal(100)) if detail else 0),
            "loai_dong": "hang_muc" if detail else "tieu_de",
            "quy_cach_model": line.get("technical_raw") if detail else None,
            "vi_tri_khu_vuc": " | ".join(str(value) for value in allocations),
            "dvt": line.get("unit_raw") if detail else None, "so_luong": qty,
            "source_row": line["source_row"], "source_stt": str(line.get("stt_raw") or ""),
            "source_item_raw": str(line.get("text_raw") or ""),
            "technical_requirement": line.get("technical_raw") if detail else None,
            "brand_raw": line.get("category_raw") if detail else None,
            "source_note_raw": line.get("note_raw") if detail else None,
        }
        insert_cols = [key for key in record if key in item_columns]
        conn.execute("INSERT INTO quotation_item(%s) VALUES(%s)" %
                     (",".join(insert_cols), ",".join("?" for _ in insert_cols)),
                     [record[key] for key in insert_cols])
        item_ids[line["source_row"]] = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    return quote_id, values["code"], item_ids


def _personnel_kind(role):
    key = _norm(role)
    if "chi huy" in key or "truong" in key:
        return "KTT"
    if "ky thuat" in key:
        return "KTV"
    return "CTV"


def _upsert_personnel(conn, project_id, import_id, parsed):
    created = linked = 0
    if parsed.get("sheet_name") is None:
        # A quote-only revision must not silently clear the current roster.
        return created, linked
    # project_personnel is the current working roster.  Rebuild it from the
    # supplied source while preserving every revision below in the snapshot.
    conn.execute("DELETE FROM project_personnel WHERE project_id=?", (project_id,))
    seen_person_ids = set()
    for row in parsed["rows"]:
        by_id = conn.execute("SELECT * FROM nhan_su WHERE cccd=?", (row["cccd"],)).fetchone() \
            if row["cccd"] else None
        by_name = [value for value in conn.execute("SELECT * FROM nhan_su").fetchall()
                   if _norm(value["ho_ten"]) == _norm(row["name"])]
        if by_id and by_name and all(value["id"] != by_id["id"] for value in by_name):
            raise ProfileImportError("CCCD va ho ten nhan su xung dot o dong %s." % row["source_row"])
        person = by_id or (by_name[0] if len(by_name) == 1 else None)
        if len(by_name) > 1 and not by_id:
            raise ProfileImportError("Trung nhieu nhan su cung ten o dong %s; can lam sach danh muc." % row["source_row"])
        if not person:
            conn.execute("""INSERT INTO nhan_su(ho_ten, loai, cccd, ngay_sinh, trang_thai,
                loai_nhan_su, ngay_vao) VALUES(?,?,?,?,?,?,?)""",
                         (row["name"], _personnel_kind(row["project_role"]), row["cccd"] or None,
                          row["birth"] or None, "Dang lam", "nha_thau_phu", date.today().isoformat()))
            person_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
            created += 1
        else:
            person_id = person["id"]
            updates, params = [], []
            for field, value in (("cccd", row["cccd"]), ("ngay_sinh", row["birth"]),
                                 ("loai_nhan_su", "nha_thau_phu")):
                if value and not person[field]:
                    updates.append(field + "=?")
                    params.append(value)
            if updates:
                conn.execute("UPDATE nhan_su SET %s WHERE id=?" % ",".join(updates), params + [person_id])
        if person_id in seen_person_ids:
            raise ProfileImportError("Nhan su bi lap trong file o dong %s." % row["source_row"])
        seen_person_ids.add(person_id)
        conn.execute("""INSERT INTO project_personnel(project_id, nhan_su_id, profile_import_id,
            source_row, source_stt, source_mst, subcontractor, site_role, project_role,
            card_expiry, card_issue_date, card_locked)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?)
            ON CONFLICT(project_id,nhan_su_id) DO UPDATE SET
              profile_import_id=excluded.profile_import_id, source_row=excluded.source_row,
              source_stt=excluded.source_stt, source_mst=excluded.source_mst,
              subcontractor=excluded.subcontractor, site_role=excluded.site_role,
              project_role=excluded.project_role, card_expiry=excluded.card_expiry,
              card_issue_date=excluded.card_issue_date, card_locked=excluded.card_locked""",
                      (project_id, person_id, import_id, row["source_row"], row["source_stt"],
                       row["source_mst"], row["subcontractor"], row["site_role"], row["project_role"],
                       row["card_expiry"], row["card_issue_date"], row["card_locked"]))
        conn.execute("""INSERT INTO project_personnel_snapshot(profile_import_id, project_id,
            nhan_su_id, source_row, source_stt, source_mst, subcontractor, site_role,
            project_role, card_expiry, card_issue_date, card_locked)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?)""",
                     (import_id, project_id, person_id, row["source_row"], row["source_stt"],
                      row["source_mst"], row["subcontractor"], row["site_role"],
                      row["project_role"], row["card_expiry"], row["card_issue_date"],
                      row["card_locked"]))
        linked += 1
    return created, linked


def _audit(conn, sess, action, table, record_id, summary):
    conn.execute("""INSERT INTO audit_log(user, role, hanh_dong, bang, ban_ghi_id, tom_tat)
                    VALUES(?,?,?,?,?,?)""",
                 (sess.get("username"), sess.get("role"), action, table,
                  str(record_id), summary[:300]))


def _seed_template_statuses(conn, project_id, template_profile=DEFAULT_TEMPLATE_PROFILE):
    try:
        import docgen
        requirements = docgen.ct_document_requirements(template_profile)
        codes = list(requirements.get("required") or ())
    except Exception:
        codes = []
    for code in codes:
        conn.execute("""INSERT INTO cong_trinh_ho_so_trang_thai(project_id, ma_mau, trang_thai)
            VALUES(?,?,'Thieu') ON CONFLICT(project_id,ma_mau) DO NOTHING""", (project_id, code))
    return len(codes)


def _generate_personnel_docs(conn, sess, project_id):
    generated, warnings = [], []
    try:
        import docgen
    except Exception as exc:
        return generated, ["Khong nap duoc docgen: %s" % exc]
    # Some focused status tests intentionally use a minimal database.  Only run
    # the production-document metadata backfill when both owning tables exist;
    # on the live schema this remains mandatory and happens before generation.
    metadata_tables = {
        row[0] for row in conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' "
            "AND name IN ('project','source_document')"
        )
    }
    project = None
    if metadata_tables == {"project", "source_document"}:
        project = conn.execute(
            "SELECT code,customer_id FROM project WHERE id=?", (project_id,)
        ).fetchone()
    for code in ("CT-01-DSNS", "CT-01-PKBNV"):
        # Backfill security metadata on older, pre-versioned generated files so
        # they cannot remain visible through the generic document index.
        if project:
            conn.execute("""UPDATE source_document SET project_id=?,profile_role='personnel'
                WHERE customer_id=? AND (project_id IS NULL OR project_id=?)
                  AND file_name LIKE ? ESCAPE '\\'
                  AND instr(file_name, ?) > 0""",
                         (project_id, project["customer_id"], project_id,
                          code.replace("_", "\\_") + "\\_%",
                          "_" + str(project["code"] or "") + "_"))
            conn.commit()
        current = conn.execute("""SELECT * FROM cong_trinh_ho_so_trang_thai
            WHERE project_id=? AND ma_mau=?""", (project_id, code)).fetchone()
        if current and current["trang_thai"] not in ("Thieu", "Dang_soan"):
            warnings.append("%s dang o trang thai %s nen khong tu sinh lai" %
                            (code, current["trang_thai"]))
            continue
        try:
            _name, _data, path = docgen.export_ct_doc(conn, sess, project_id, code)
            if not path:
                conn.rollback()
                warnings.append("%s: khong luu duoc file vao ho so cong trinh" % code)
                continue
            if current:
                conn.execute("""UPDATE cong_trinh_ho_so_trang_thai SET trang_thai='Dang_soan',
                    file_path=?, updated_by=?, updated_at=datetime('now') WHERE id=?""",
                             (path, sess.get("user_id"), current["id"]))
            else:
                conn.execute("""INSERT INTO cong_trinh_ho_so_trang_thai(project_id,ma_mau,
                    trang_thai,file_path,updated_by) VALUES(?,?,'Dang_soan',?,?)""",
                             (project_id, code, path, sess.get("user_id")))
            conn.commit()
            generated.append(code)
        except Exception as exc:
            conn.rollback()
            warnings.append("%s: %s" % (code, exc))
    return generated, warnings


def commit_project_profile(conn, sess, data):
    """Commit a previously previewed profile.  Same bundle hash is a no-op."""
    _require_role(sess)
    token = (data.get("confirm_token") or "").strip()
    _cleanup_previews()
    with _PREVIEW_LOCK:
        state = _PREVIEWS.get(token)
    if not state:
        raise ProfileImportError("Preview da het han/khong ton tai; hay xem truoc lai.")
    audit = state["quote"].get("normalization_audit") or {}
    if state["quote"].get("document_kind") != "QUOTATION":
        raise ProfileImportError(
            "File da nhan dien la %s, khong duoc ghi thanh bao gia chinh thuc. "
            "Hay import tai dung nghiep vu phat sinh/thanh toan." %
            state["quote"].get("document_kind"))
    if not audit.get("ready_for_official_commit"):
        raise ProfileImportError(
            "Sandbox con %d sai lech chan ghi; phai sua va Preview lai truoc khi commit." %
            len(audit.get("blocking_issues") or []))
    _verify_files_unchanged(state)
    backup_path = _backup_database(conn)
    conn.execute("BEGIN IMMEDIATE")
    materialized_uploads = []
    try:
        project = conn.execute("SELECT * FROM project WHERE id=?", (state["project_id"],)).fetchone() \
            if state["project_id"] else None
        if project:
            project_id = project["id"]
            customer_id = project["customer_id"]
        else:
            customer_id = state["customer_id"]
            code = _next_code(conn, "project", "CT")
            conn.execute("""INSERT INTO project(code,project_name,customer_id,status,percent_complete,
                                template_profile) VALUES(?,?,?,'Open',0,?)""",
                         (code, state["project_name"], customer_id, state["template_profile"]))
            project_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        if project and "template_profile" in project.keys() and \
                project["template_profile"] != state["template_profile"]:
            conn.execute("UPDATE project SET template_profile=? WHERE id=?",
                         (state["template_profile"], project_id))

        duplicate = conn.execute("""SELECT id, quotation_id FROM project_profile_import
            WHERE project_id=? AND bundle_sha256=?""", (project_id, state["bundle_sha256"])).fetchone()
        if duplicate:
            conn.rollback()
            with _PREVIEW_LOCK:
                _PREVIEWS.pop(token, None)
            return {"ok": True, "idempotent": True, "project_id": project_id,
                    "profile_import_id": duplicate["id"], "quotation_id": duplicate["quotation_id"]}

        active = conn.execute("""SELECT * FROM project_profile_import
            WHERE project_id=? AND status='active' ORDER BY id DESC LIMIT 1""", (project_id,)).fetchone()
        if active:
            used = conn.execute("""SELECT COUNT(*) FROM project_boq_stage_qty q
                JOIN project_boq_line l ON l.id=q.boq_line_id
                WHERE l.profile_import_id=? AND
                (q.actual_qty<>0 OR q.returned_qty<>0 OR EXISTS(
                    SELECT 1 FROM stock_ledger s WHERE s.boq_stage_qty_id=q.id))""",
                                (active["id"],)).fetchone()[0]
            if used:
                raise ProfileImportError("BOQ hien tai da co thuc te/xuat kho; khong tu thay the. Can quy trinh revision co doi chieu.")

        # Chi chan khi DINH MUC cu da co thuc te/hoan tra (se bi xoa/thay).
        # Phieu xuat kho chua gan boq_stage_qty_id la lich su cu (truoc BOQ) —
        # KHONG duoc chan revision bao gia/BOQ ke hoach khi BOQ active chua co actual
        # (da kiem o `used` phia tren). Neu chan o day, KPI du toan khong bao gio cap nhat.
        legacy_used = conn.execute("""SELECT COUNT(*) FROM cong_trinh_dinh_muc_vat_tu
            WHERE project_id=? AND (kl_thuc_te<>0 OR kl_hoan_tra<>0)""", (project_id,)).fetchone()[0]
        if legacy_used:
            raise ProfileImportError(
                "Dinh muc cu da co khoi luong thuc te/hoan tra; khong duoc tu xoa/thay. "
                "Can quy trinh revision co doi chieu.")
        legacy_stock_used = conn.execute("""SELECT COUNT(*) FROM stock_ledger
            WHERE project_id=? AND lower(movement_type)='xuat_cong_trinh'
              AND COALESCE(qty_out,0)<>0 AND boq_stage_qty_id IS NULL""",
                                         (project_id,)).fetchone()[0]
        if legacy_stock_used:
            # Canh bao — van cho commit revision BG/BOQ ke hoach.
            state.setdefault("warning_metadata", []).append({
                "code": "LEGACY_STOCK_UNLINKED_TO_BOQ",
                "count": int(legacy_stock_used),
                "message": (
                    "%d dong xuat kho chua gan BOQ (phieu cu) — giu nguyen so cai; "
                    "chi cap nhat bao gia/BOQ ke hoach." % int(legacy_stock_used)),
            })

        for info, folder in ((state["quote_file"], "Báo giá"),
                             (state["contract_file"], "Hợp đồng"),
                             (state["personnel_file"], "Hồ sơ công trình")):
            created = _materialize_upload(conn, customer_id, state, info, folder)
            if created:
                materialized_uploads.append(created)
        quote_doc = _index_source(conn, customer_id, project_id, state["quote_file"],
                                  "official_quote", "Bao gia")
        contract_doc = _index_source(conn, customer_id, project_id, state["contract_file"],
                                     "contract", "Hop dong") if state["contract_file"] else None
        personnel_doc = _index_source(conn, customer_id, project_id, state["personnel_file"],
                                      "personnel", "Ho so") if state["personnel_file"] else None

        current_quote = _current_quote(conn, project_id)
        quote_id, quote_code, item_ids = _insert_official_quote(
            conn, project_id, customer_id, current_quote, state)
        contract = _existing_contract(conn, project_id, state["contract_id"])
        contract_id = contract["id"] if contract else None
        if contract:
            conn.execute("UPDATE hop_dong_ct SET quotation_id=? WHERE id=?", (quote_id, contract_id))

        if active:
            conn.execute("UPDATE project_profile_import SET status='superseded' WHERE id=?", (active["id"],))
        warning_codes = Counter(item.get("code", "OTHER") for item in
                                (state.get("warning_metadata") or []))
        normalization_audit = state["quote"].get("normalization_audit") or {}
        conn.execute("""INSERT INTO project_profile_import(project_id,quotation_id,contract_id,
            quote_document_id,contract_document_id,personnel_document_id,source_file_name,
            source_sha256,contract_sha256,personnel_sha256,bundle_sha256,source_sheet,
            parser_version,status,detail_count,heading_count,stage_count,warning_json,
            normalization_version,normalization_status,normalization_audit_json,
            source_amount_total,normalized_amount_total,money_tolerance_ratio,imported_by)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,'active',?,?,?,?,?,?,?,?,?,?,?)""",
                     (project_id, quote_id, contract_id, quote_doc, contract_doc, personnel_doc,
                      state["quote_file"]["filename"], state["quote_file"]["sha256"],
                      state["contract_file"]["sha256"] if state["contract_file"] else None,
                      state["personnel_file"]["sha256"] if state["personnel_file"] else None,
                      state["bundle_sha256"], state["quote"]["source"]["sheet_name"],
                      state["quote"].get("parser_version"), state["quote"]["counts"]["detail_count"],
                      state["quote"]["counts"]["heading_count"], len(state["quote"]["stages"]),
                      json.dumps(dict(warning_codes), ensure_ascii=False),
                      state["quote"].get("parser_version"), "PASSED",
                      json.dumps(normalization_audit, ensure_ascii=False),
                      _num(normalization_audit.get("source_amount_total")),
                      _num(normalization_audit.get("normalized_amount_total")),
                      _num(normalization_audit.get("money_tolerance_ratio")), sess.get("username")))
        import_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]

        for row_check in normalization_audit.get("row_checks") or []:
            fields = row_check.get("fields") or {}
            source_values = {name: value.get("source") for name, value in fields.items()}
            normalized_values = {name: value.get("normalized") for name, value in fields.items()}
            conn.execute("""INSERT INTO project_import_normalization_audit(profile_import_id,
                source_sheet,source_row,item_name,source_values_json,normalized_values_json,result_json)
                VALUES(?,?,?,?,?,?,?)""", (import_id,
                row_check.get("source_sheet") or state["quote"]["source"]["sheet_name"],
                int(row_check["source_row"]), row_check.get("item_name"),
                json.dumps(source_values, ensure_ascii=False),
                json.dumps(normalized_values, ensure_ascii=False),
                json.dumps({"fields": fields, "stage_status": row_check.get("stage_status"),
                            "classification_confidence": row_check.get("classification_confidence")},
                           ensure_ascii=False)))

        stage_ids = {}
        for stage in state["quote"]["stages"]:
            source_index = stage.get("source_index")
            source_col = int(source_index) + 1 if source_index is not None else None
            conn.execute("""INSERT INTO project_boq_stage(profile_import_id,thu_tu,source_col,
                name_raw,name_normalized,is_unallocated) VALUES(?,?,?,?,?,?)""",
                         (import_id, stage["order"], source_col, str(stage["name"]),
                           _norm(stage["name"]), 1 if stage["is_unallocated_bucket"] else 0))
            stage_ids[stage["key"]] = conn.execute("SELECT last_insert_rowid()").fetchone()[0]

        allocation_count = 0
        for line in state["quote"]["lines"]:
            detail = line["kind"] == "detail"
            floor_total = line.get("stage_total") or {}
            contract_qty = line.get("contract_quantity") or {}
            unit_price = line.get("unit_price") or {}
            amount = line.get("amount") or {}
            vat = _vat_for_line(line) if detail else None
            conn.execute("""INSERT INTO project_boq_line(profile_import_id,quotation_item_id,
                source_sheet,source_row,thu_tu,line_type,hierarchy_level,hierarchy_path,
                source_stt_raw,item_name_raw,technical_requirement_raw,uom_raw,
                floor_total_qty_raw,floor_total_qty,contract_qty_raw,contract_qty,
                unit_price_raw,unit_price,amount_raw,amount,brand_raw,note_raw,vat_rate_raw)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                         (import_id, item_ids[line["source_row"]], state["quote"]["source"]["sheet_name"],
                          line["source_row"], line["source_order"], line["kind"],
                          line.get("heading_level", len(line.get("hierarchy_path") or [])),
                          json.dumps(line.get("hierarchy_path") or [], ensure_ascii=False),
                          str(line.get("stt_raw") or ""), str(line.get("text_raw") or ""),
                          line.get("technical_raw"), line.get("unit_raw"), _raw(floor_total),
                          _num(floor_total.get("value")), _raw(contract_qty), _num(contract_qty.get("value")),
                          _raw(unit_price), _num(unit_price.get("value")), _raw(amount),
                          _num(amount.get("value")), line.get("category_raw"), line.get("note_raw"),
                          str(vat) if vat is not None else None))
            line_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
            if not detail:
                continue
            quantities = line.get("stage_quantities") or []
            if quantities:
                for quantity in quantities:
                    conn.execute("""INSERT INTO project_boq_stage_qty(boq_line_id,stage_id,
                        planned_qty,planned_qty_raw) VALUES(?,?,?,?)""",
                                 (line_id, stage_ids[quantity["stage_key"]],
                                  _num(quantity.get("value")) or 0, _raw(quantity) or quantity.get("value") or "0"))
                    allocation_count += 1
            else:
                planned = _num(floor_total.get("value")) or 0
                conn.execute("""INSERT INTO project_boq_stage_qty(boq_line_id,stage_id,
                    planned_qty,planned_qty_raw) VALUES(?,?,?,?)""",
                             (line_id, stage_ids["unallocated"], planned,
                              _raw(floor_total) or floor_total.get("value") or "0"))
                allocation_count += 1

        created_people, linked_people = _upsert_personnel(
            conn, project_id, import_id, state["personnel"])
        legacy_removed = conn.execute(
            "DELETE FROM cong_trinh_dinh_muc_vat_tu WHERE project_id=?", (project_id,)).rowcount
        template_count = _seed_template_statuses(conn, project_id, state["template_profile"])
        _audit(conn, sess, "PROJECT_PROFILE_IMPORT", "project_profile_import", import_id,
               "Import official profile project=%s quote=%s lines=%d stages=%d personnel=%d; preserve history" %
               (project_id, quote_code, state["quote"]["counts"]["line_count"],
                len(state["quote"]["stages"]), linked_people))
        conn.commit()
    except Exception:
        conn.rollback()
        _cleanup_materialized_uploads(materialized_uploads)
        raise

    generated, generation_warnings = ([], [])
    if state["auto_generate_templates"] and state["personnel"]["rows"]:
        generated, generation_warnings = _generate_personnel_docs(conn, sess, project_id)
    with _PREVIEW_LOCK:
        _PREVIEWS.pop(token, None)
    return {
        "ok": True, "idempotent": False, "project_id": project_id,
        "profile_import_id": import_id, "quotation_id": quote_id,
        "quotation_code": quote_code,
        "counts": {
            "headings": state["quote"]["counts"]["heading_count"],
            "details": state["quote"]["counts"]["detail_count"],
            "stages_including_unallocated": len(state["quote"]["stages"]),
            "stage_allocations": allocation_count,
            "personnel_linked": linked_people, "personnel_created": created_people,
            "legacy_predicted_rows_removed": legacy_removed,
            "templates_registered": template_count,
        },
        "generated_templates": generated,
        "generation_warnings": generation_warnings,
        "backup_created": bool(backup_path),
        "preserved": ["project", "contract", "workflow", "existing_dossier_history", "quotation_history"],
    }


def update_boq_actual(conn, sess, data):
    if (sess or {}).get("role") not in ("Giam doc", "Ky thuat truong", "Quan tri he thong"):
        raise ProfileImportError("Vai tro hien tai khong co quyen cap nhat khoi luong BOQ.")
    try:
        row_id = int(data.get("id"))
        actual = float(data.get("actual_qty") or 0)
        returned = float(data.get("returned_qty") or 0)
    except (TypeError, ValueError):
        raise ProfileImportError("Khoi luong thuc te/hoan tra phai la so.")
    if actual < 0 or returned < 0:
        raise ProfileImportError("Khoi luong khong duoc am.")
    status = (data.get("status") or "Chua_doi_chieu").strip()
    allowed = {"Chua_doi_chieu", "Khop", "Cho_xac_nhan", "Cho_doi_chieu", "Vuot_du_toan"}
    if status not in allowed:
        raise ProfileImportError("Trang thai doi chieu khong hop le.")
    conn.execute("BEGIN IMMEDIATE")
    try:
        row = conn.execute("""SELECT q.*, i.project_id FROM project_boq_stage_qty q
            JOIN project_boq_line l ON l.id=q.boq_line_id
            JOIN project_profile_import i ON i.id=l.profile_import_id
            WHERE q.id=? AND i.status='active'""", (row_id,)).fetchone()
        if not row:
            raise ProfileImportError("Dong BOQ khong ton tai/khong con active.")
        expected = data.get("expected_updated_at")
        if expected is not None and str(expected or "") != str(row["updated_at"] or ""):
            raise ProfileImportError("Dong BOQ da duoc nguoi khac cap nhat; hay tai lai truoc khi ghi.")
        conn.execute("""INSERT INTO project_boq_actual_log(stage_qty_id,actual_qty_before,
            actual_qty_after,returned_qty_before,returned_qty_after,status_before,status_after,note,changed_by)
            VALUES(?,?,?,?,?,?,?,?,?)""",
                     (row_id, row["actual_qty"], actual, row["returned_qty"], returned,
                      row["status"], status, data.get("note"), sess.get("username")))
        conn.execute("""UPDATE project_boq_stage_qty SET actual_qty=?,returned_qty=?,status=?,note=?,
            updated_at=strftime('%Y-%m-%dT%H:%M:%fZ','now') WHERE id=?""",
                     (actual, returned, status, data.get("note"), row_id))
        _audit(conn, sess, "PROJECT_BOQ_ACTUAL", "project_boq_stage_qty", row_id,
               "Update actual project=%s" % row["project_id"])
        updated_at = conn.execute(
            "SELECT updated_at FROM project_boq_stage_qty WHERE id=?", (row_id,)).fetchone()[0]
        conn.commit()
        return {"ok": True, "id": row_id, "updated_at": updated_at}
    except Exception:
        conn.rollback()
        raise
